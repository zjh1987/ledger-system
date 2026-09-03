# -*- coding: utf-8 -*-
"""
采购代理台账管理系统 - Flask 主应用
提供 10 个台账表的 REST API，支持:
  - CRUD 操作 (GET / POST / PUT / DELETE)
  - 分页查询 (?page=1&per_page=20)
  - 搜索 (?search=keyword)
  - 数据校验 (/api/<table_name>/validate)
  - Excel 导出 (/api/export/<table_name>)
  - Excel 导入 (/api/import)
  - 表结构查询 (/api/tables, /api/columns/<table_name>)

路由分组索引 (用于快速跳转):
  - L~537    页面路由 (/, /static/, /analytics)
  - L~541    认证 (login/logout/me)
  - L~620    用户管理 (users CRUD, 仅管理员)
  - L~754    表结构 (/api/tables, /api/eval_rooms/calc)
  - L~810    列管理 (/api/columns/*)
  - L~1257   下拉选项管理 (/api/dropdowns/*)
  - L~1415   自定义列管理 (/api/custom-columns/*)
  - L~1669   通用 CRUD (/api/<table>, /api/<table>/<id>, /validate, /batch)
  - L~2085   Excel 导出 (/api/export, /export-all)
  - L~2307   Excel 导入 (/api/import, /import-all)
  - L~2729   导入模板 (/api/import-template)
  - L~2917   统计与运维 (/api/stats, /health, /clear-all)
  - L~3003   数据分析 (/api/analytics/*)
  - L~3417   主入口 (ensure_tables + app.run)
"""

import os
import io
import json
import time
import sqlite3
import secrets
import hashlib
import re
from datetime import datetime
from functools import wraps

from flask import Flask, request, jsonify, send_file, Response, render_template, send_from_directory, redirect
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from werkzeug.security import generate_password_hash, check_password_hash

from schema import (
    TABLES,
    get_fields,
    get_field_keys,
    get_field_cn_names,
    get_field_by_key,
    get_primary_key,
    get_create_table_sql,
    get_all_create_table_sql,
    get_dropdown_fields,
    get_table_info,
    get_all_tables_info,
    get_table_cn_name,
    get_sqlite_type,
    get_visible_fields,
    get_seq_fields,
    is_internal_table,
    get_user_tables,
)
from validator import validate_record, validate_batch, validate_table, validate_field, _get_merged_dropdown_options, DB_PATH as VALIDATOR_DB_PATH
from formula_calc import compute_formula_fields, should_recalc_on_update, get_dependent_eval_rooms
from utils import (
    BEIJING_TZ,
    now_beijing,
    parse_date,
    parse_number,
    to_int,
    serialize_value,
    row_to_dict,
    clean_record,
)

# ============================================================
# 配置
# ============================================================

# 数据库路径: 优先环境变量, 默认当前目录
_default_db = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tz.db")
DB_PATH = os.environ.get("DB_PATH", _default_db)

# 确保数据库目录存在
_db_dir = os.path.dirname(DB_PATH)
if _db_dir and not os.path.exists(_db_dir):
    try:
        os.makedirs(_db_dir, exist_ok=True)
    except (OSError, PermissionError):
        DB_PATH = _default_db

# 同步 validator 的数据库路径
import validator as _validator_module
_validator_module.DB_PATH = DB_PATH


# ============================================================
# Flask 应用初始化
# ============================================================
app = Flask(__name__)
# 默认生产环境关闭 debug；本地开发可通过 FLASK_DEBUG=1 开启
app.config["DEBUG"] = os.environ.get("FLASK_DEBUG", "0") == "1"
# SECRET_KEY: 用于 Flask session 等；如需使用可从环境变量注入
app.config["SECRET_KEY"] = os.environ.get("FLASK_SECRET_KEY", secrets.token_hex(32))
# CORS: 默认仅允许同源请求（不额外发送 CORS 头）。确需跨域时通过 ALLOWED_ORIGINS 逗号分隔配置白名单。
_allowed_origins = [o.strip() for o in os.environ.get("ALLOWED_ORIGINS", "").split(",") if o.strip()]
CORS(app, resources={r"/api/*": {"origins": _allowed_origins}})


# ============================================================
# 数据库辅助函数
# ============================================================

def get_db():
    """
    获取数据库连接。
    自动启用 SQLite WAL 模式 + busy_timeout，提高并发读写的健壮性。
    WAL 模式在容器/重启/旧版本上若失败则自动回退到默认 rollback 模式。
    """
    conn = sqlite3.connect(DB_PATH, timeout=5.0)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        conn.execute("PRAGMA journal_mode = WAL")
    except sqlite3.DatabaseError:
        # 老版本 SQLite 不支持 WAL，自动忽略
        pass
    try:
        conn.execute("PRAGMA busy_timeout = 5000")
    except sqlite3.DatabaseError:
        pass
    return conn


def ensure_tables():
    """确保所有表已创建"""
    conn = get_db()
    cursor = conn.cursor()
    for table_name, sql in get_all_create_table_sql().items():
        cursor.execute(sql)

    # 创建用户表
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            display_name TEXT,
            role TEXT DEFAULT 'user',
            must_change_password INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now', 'localtime'))
        )
    """)

    # 兼容旧库：为 users 表补齐 must_change_password 列
    cursor.execute('PRAGMA table_info("users")')
    users_columns = [row[1] for row in cursor.fetchall()]
    if "must_change_password" not in users_columns:
        cursor.execute('ALTER TABLE "users" ADD COLUMN must_change_password INTEGER DEFAULT 0')

    # 插入默认管理员账号 (如不存在)。初始密码为默认值，强制首次修改。
    cursor.execute("SELECT COUNT(*) FROM users WHERE username = ?", ("admin",))
    if cursor.fetchone()[0] == 0:
        cursor.execute(
            "INSERT INTO users (username, password, display_name, role, must_change_password) VALUES (?, ?, ?, ?, 1)",
            ("admin", hash_password("admin123"), "管理员", "admin"),
        )

    # 检查默认管理员是否仍使用默认密码（admin123），用于启动时警告
    cursor.execute("SELECT password FROM users WHERE username = ?", ("admin",))
    row = cursor.fetchone()
    if row and verify_password(row["password"], "admin123"):
        # 仅置标志位，警告在主入口打印（避免重复）
        app.config["DEFAULT_PASSWORD_WARNING"] = True

    # 给所有业务表添加 created_by 字段 (如不存在)
    for table_name in TABLES.keys():
        cursor.execute(f'PRAGMA table_info("{table_name}")')
        columns = [row[1] for row in cursor.fetchall()]
        if "created_by" not in columns:
            cursor.execute(f'ALTER TABLE "{table_name}" ADD COLUMN created_by TEXT')
        # 给所有业务表添加 updated_at 字段 (如不存在) —— 记录最近修改时间
        if "updated_at" not in columns:
            cursor.execute(
                f'ALTER TABLE "{table_name}" ADD COLUMN updated_at TEXT'
            )
            # 为已有记录填充当前北京时间
            cursor.execute(
                f'UPDATE "{table_name}" SET updated_at = datetime("now", "+8 hours") '
                f'WHERE updated_at IS NULL'
            )

    # 给 _custom_columns 表添加 after_field 字段 (如不存在)
    cursor.execute('PRAGMA table_info("_custom_columns")')
    cc_columns = [row[1] for row in cursor.fetchall()]
    if "after_field" not in cc_columns:
        cursor.execute('ALTER TABLE "_custom_columns" ADD COLUMN after_field TEXT')

    # 创建表头重命名表
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS _column_renames (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            table_name TEXT NOT NULL,
            field_key TEXT NOT NULL,
            custom_name TEXT NOT NULL,
            original_name TEXT,
            updated_at TEXT DEFAULT (datetime('now', 'localtime')),
            UNIQUE(table_name, field_key)
        )
    """)

    # 创建下拉选项管理表
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS dropdown_options (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            table_name TEXT NOT NULL,
            field_key TEXT NOT NULL,
            option_value TEXT NOT NULL,
            sort_order INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now', 'localtime')),
            UNIQUE(table_name, field_key, option_value)
        )
    """)

    # 如果 dropdown_options 表为空, 从 schema 初始化默认选项
    cursor.execute("SELECT COUNT(*) FROM dropdown_options")
    if cursor.fetchone()[0] == 0:
        for t_name, t_info in TABLES.items():
            # 跳过内部表
            if t_info.get("internal"):
                continue
            for field in t_info["fields"]:
                options = field.get("options")
                if not options:
                    continue
                for idx, opt in enumerate(options):
                    cursor.execute(
                        "INSERT OR IGNORE INTO dropdown_options "
                        "(table_name, field_key, option_value, sort_order) VALUES (?, ?, ?, ?)",
                        (t_name, field["key"], opt, idx),
                    )

    # 为含可编辑序号的表添加 (序号字段, created_by) 唯一索引：
    # 业务规则 = 同一账号可复用同一序号，不同账号不能使用相同序号（SQLite 唯一索引天然满足）。
    for t_name, t_info in get_user_tables().items():
        seq_fields = get_seq_fields(t_name)
        if seq_fields:
            seq_col = seq_fields[0]["key"]
            idx_name = f"idx_{t_name}_{seq_col}_owner"
            try:
                cursor.execute(
                    f'CREATE UNIQUE INDEX IF NOT EXISTS "{idx_name}" '
                    f'ON "{t_name}" ("{seq_col}", "created_by")'
                )
            except sqlite3.Error:
                # 已有历史重复数据导致建索引失败时，仅跳过（应用层仍会校验）
                pass

    conn.commit()
    conn.close()


def get_all_field_keys_with_custom(table_name):
    """获取表的所有字段key，包括自定义列（用于 CRUD 过滤）"""
    keys = list(get_field_keys(table_name))
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT field_key FROM _custom_columns WHERE table_name = ? ORDER BY sort_order, id",
            (table_name,),
        )
        for row in cursor.fetchall():
            fk = row["field_key"]
            if fk not in keys:
                keys.append(fk)
        conn.close()
    except sqlite3.Error:
        pass
    return keys


def get_hidden_field_keys(table_name):
    """获取指定表中已隐藏的基础列 field_key 集合"""
    hidden_keys = set()
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT field_key FROM _hidden_columns WHERE table_name = ?",
            (table_name,),
        )
        hidden_keys = {row["field_key"] for row in cursor.fetchall()}
        conn.close()
    except sqlite3.Error:
        pass
    return hidden_keys


def get_column_renames(table_name):
    """获取指定表中已自定义重命名的字段 {field_key: custom_name}"""
    renames = {}
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT field_key, custom_name FROM _column_renames WHERE table_name = ?",
            (table_name,),
        )
        for row in cursor.fetchall():
            renames[row["field_key"]] = row["custom_name"]
        conn.close()
    except sqlite3.Error:
        pass
    return renames


def apply_column_renames(table_name, fields):
    """将自定义重命名应用到字段列表（原地修改 name_cn）"""
    renames = get_column_renames(table_name)
    if not renames:
        return fields
    for f in fields:
        if f.get("key") in renames:
            f["original_name_cn"] = f.get("name_cn", "")
            f["name_cn"] = renames[f["key"]]
            f["is_renamed"] = True
    return fields


def get_visible_fields_with_custom(table_name):
    """
    获取可见字段列表（schema 可见字段 + 自定义列），用于导出/导入/列定义。
    - 排除 _hidden_columns 中隐藏的基础列
    - 按 sort_order 排序所有列（基础列和自定义列混合排序）
    - 基础列的默认 sort_order 按其在 schema 中的位置（0, 1, 2...）
    - 自定义列按其在 _custom_columns 中的 sort_order，并参考 after_field 决定插入位置
    """
    # 1. 获取 schema 可见字段（排除 hidden 字段如 id）
    base_fields = list(get_visible_fields(table_name))

    # 2. 排除被隐藏的基础列
    hidden_keys = get_hidden_field_keys(table_name)
    visible_base_fields = [f for f in base_fields if f["key"] not in hidden_keys]

    # 3. 获取自定义列
    custom_fields = []
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT * FROM _custom_columns WHERE table_name = ? ORDER BY sort_order, id",
            (table_name,),
        )
        for row in cursor.fetchall():
            after_field = row["after_field"] if "after_field" in row.keys() else None
            custom_fields.append({
                "key": row["field_key"],
                "name_cn": row["field_name_cn"],
                "type": row["field_type"],
                "required": False,
                "is_custom": True,
                "is_base": False,
                "after_field": after_field,
                "sort_order": row["sort_order"],
            })
        conn.close()
    except sqlite3.Error:
        pass

    # 4. 构建有序结果列表
    #    基础列按 schema 顺序排列，自定义列按 after_field 插入到对应基础列之后
    result = []
    placed_custom_keys = set()

    # 第一轮：遍历基础列，将 after_field 指向当前基础列的自定义列插入其后
    for idx, bf in enumerate(visible_base_fields):
        field_copy = dict(bf)
        field_copy["is_base"] = True
        field_copy["is_custom"] = False
        field_copy["is_hidden"] = bf["key"] in hidden_keys
        field_copy["sort_order"] = idx
        result.append(field_copy)

        # 查找 after_field == 当前基础列 key 的自定义列（已按 sort_order 排序）
        customs_after = [
            cf for cf in custom_fields
            if cf.get("after_field") == bf["key"] and cf["key"] not in placed_custom_keys
        ]
        for cf in customs_after:
            result.append(cf)
            placed_custom_keys.add(cf["key"])

    # 第二轮：处理 after_field 指向其他自定义列的情况（链式插入）
    changed = True
    max_iterations = len(custom_fields) + 1
    iteration = 0
    while changed and iteration < max_iterations:
        changed = False
        iteration += 1
        for cf in custom_fields:
            if cf["key"] in placed_custom_keys:
                continue
            after = cf.get("after_field")
            if after:
                # 在已放置的结果中查找 after_field 的位置
                for i, item in enumerate(result):
                    if item["key"] == after:
                        result.insert(i + 1, cf)
                        placed_custom_keys.add(cf["key"])
                        changed = True
                        break

    # 第三轮：将未放置的自定义列（无 after_field 或 after_field 不存在）追加到末尾
    for cf in custom_fields:
        if cf["key"] not in placed_custom_keys:
            result.append(cf)
            placed_custom_keys.add(cf["key"])

    # 应用自定义重命名
    apply_column_renames(table_name, result)

    return result


# ============================================================
# 装饰器: 检查表是否存在
# ============================================================

def table_exists_decorator(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        table_name = kwargs.get("table_name")
        if table_name and table_name not in TABLES:
            return jsonify({
                "success": False,
                "error": f"未知的表名: {table_name}",
                "available_tables": list(get_user_tables().keys()),
            }), 404
        if table_name and is_internal_table(table_name):
            return jsonify({
                "success": False,
                "error": f"无权访问内部表: {table_name}",
            }), 403
        return f(*args, **kwargs)
    return decorated


# ============================================================
# 用户认证与权限控制
# ============================================================

# token 存储: {token: {"username": "xxx", "role": "admin/user", "display_name": "xxx", "user_id": int, "expires_at": float}}
TOKENS = {}
TOKEN_EXPIRE_SECONDS = 86400  # 24 小时过期

# 登录失败频率限制：{key: {"count": int, "first_attempt": float}}
_LOGIN_FAILURES = {}
_LOGIN_MAX_ATTEMPTS = 5       # 5 次失败后锁定
_LOGIN_LOCKOUT_MINUTES = 15   # 锁定 15 分钟


def _login_rate_limit_key(username, ip):
    """生成限流 key（按用户名+IP 组合）"""
    return f"{username}:{ip}"


def _check_login_rate_limit(username, ip):
    """检查登录是否被限流。返回剩余尝试次数，0 表示已锁定。"""
    key = _login_rate_limit_key(username, ip)
    now = time.time()
    entry = _LOGIN_FAILURES.get(key)
    if entry:
        # 超过锁定时间则重置
        if now - entry["first_attempt"] > _LOGIN_LOCKOUT_MINUTES * 60:
            _LOGIN_FAILURES.pop(key, None)
            return _LOGIN_MAX_ATTEMPTS
        if entry["count"] >= _LOGIN_MAX_ATTEMPTS:
            return 0
        return _LOGIN_MAX_ATTEMPTS - entry["count"]
    return _LOGIN_MAX_ATTEMPTS


def _record_login_failure(username, ip):
    """记录一次登录失败。"""
    key = _login_rate_limit_key(username, ip)
    now = time.time()
    entry = _LOGIN_FAILURES.get(key)
    if entry:
        entry["count"] += 1
    else:
        _LOGIN_FAILURES[key] = {"count": 1, "first_attempt": now}


def _clear_login_rate_limit(username, ip):
    """登录成功后清除限流记录。"""
    key = _login_rate_limit_key(username, ip)
    _LOGIN_FAILURES.pop(key, None)


def hash_password(password):
    """
    对密码进行哈希。
    使用 werkzeug PBKDF2-SHA256（600k 轮 + 自动 salt），格式如:
        pbkdf2:sha256:600000$<salt>$<hash>
    """
    return generate_password_hash(password, method="pbkdf2:sha256", salt_length=16)


def _is_legacy_sha256_hash(stored: str) -> bool:
    """检测是否为旧版 SHA-256 无盐哈希（64位小写hex）。"""
    return bool(stored) and bool(re.fullmatch(r"[0-9a-f]{64}", stored))


def verify_password(stored: str, candidate: str) -> bool:
    """
    校验密码。支持两种存储格式：
      1) werkzeug PBKDF2 哈希（当前默认）
      2) 旧版 SHA-256 无盐哈希（仅做一次性兼容，验证通过后调用方应重写）
    """
    if not stored or not candidate:
        return False
    if _is_legacy_sha256_hash(stored):
        return hashlib.sha256(candidate.encode()).hexdigest() == stored
    try:
        return check_password_hash(stored, candidate)
    except (ValueError, TypeError):
        return False


def create_token(username, role, display_name, user_id):
    """生成随机 token 并存入内存字典（带过期时间）"""
    token = secrets.token_hex(32)
    TOKENS[token] = {
        "username": username,
        "role": role,
        "display_name": display_name,
        "user_id": user_id,
        "expires_at": time.time() + TOKEN_EXPIRE_SECONDS,
    }
    return token


def _clean_expired_tokens():
    """清理过期 token。"""
    now = time.time()
    expired = [t for t, v in TOKENS.items() if v.get("expires_at", 0) < now]
    for t in expired:
        TOKENS.pop(t, None)


def get_current_user():
    """从请求头解析 token 获取当前用户信息。自动清理过期 token。"""
    _clean_expired_tokens()
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        token = auth[7:]
        data = TOKENS.get(token)
        if data and data.get("expires_at", 0) > time.time():
            return data
        # 过期后清除
        if data:
            TOKENS.pop(token, None)
    return None


def require_auth(f):
    """要求用户已登录"""
    @wraps(f)
    def decorated(*args, **kwargs):
        user = get_current_user()
        if not user:
            return jsonify({"success": False, "error": "未登录或登录已过期，请重新登录"}), 401
        request.current_user = user
        return f(*args, **kwargs)
    return decorated


def require_admin(f):
    """要求用户已登录且为管理员"""
    @wraps(f)
    def decorated(*args, **kwargs):
        user = get_current_user()
        if not user:
            return jsonify({"success": False, "error": "未登录或登录已过期"}), 401
        if user["role"] != "admin":
            return jsonify({"success": False, "error": "需要管理员权限"}), 403
        request.current_user = user
        return f(*args, **kwargs)
    return decorated


# ============================================================
# 页面路由
# ============================================================

@app.route("/")
def index():
    """主页面"""
    return render_template("index.html")


@app.route("/static/<path:filename>")
def serve_static(filename):
    """静态文件"""
    return send_from_directory("static", filename)


# ============================================================
# API: 用户认证 (登录/登出/当前用户)
# ============================================================

@app.route("/api/login", methods=["POST"])
def login():
    """用户登录, 返回 token"""
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({"success": False, "error": "请求体必须为 JSON 格式"}), 400

    username = (data.get("username") or "").strip()
    password = data.get("password") or ""

    if not username or not password:
        return jsonify({"success": False, "error": "用户名和密码不能为空"}), 400

    client_ip = request.remote_addr or "unknown"

    # 登录限流：同一用户名+IP 连续失败多次后锁定
    remaining = _check_login_rate_limit(username, client_ip)
    if remaining == 0:
        return jsonify({
            "success": False,
            "error": f"登录失败次数过多，请 {_LOGIN_LOCKOUT_MINUTES} 分钟后再试",
        }), 429

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT * FROM users WHERE username = ?",
        (username,),
    )
    user = cursor.fetchone()

    if not user:
        _record_login_failure(username, client_ip)
        conn.close()
        return jsonify({"success": False, "error": "用户名或密码错误"}), 401

    # 校验密码（兼容旧 SHA-256 哈希，验证通过后自动升级为 PBKDF2）
    if not verify_password(user["password"], password):
        _record_login_failure(username, client_ip)
        conn.close()
        return jsonify({"success": False, "error": "用户名或密码错误"}), 401

    # 登录成功，清除限流记录
    _clear_login_rate_limit(username, client_ip)

    # 自动迁移：旧格式哈希升级为新格式（一次性）
    if _is_legacy_sha256_hash(user["password"]):
        try:
            cursor.execute(
                "UPDATE users SET password = ? WHERE id = ?",
                (hash_password(password), user["id"]),
            )
            conn.commit()
        except sqlite3.Error:
            # 迁移失败不影响本次登录
            pass

    display_name = user["display_name"] or user["username"]
    must_change_password = bool(user["must_change_password"]) if "must_change_password" in user.keys() else False
    token = create_token(user["username"], user["role"], display_name, user["id"])
    conn.close()
    return jsonify({
        "success": True,
        "token": token,
        "user": {
            "username": user["username"],
            "role": user["role"],
            "display_name": display_name,
            "id": user["id"],
            "must_change_password": must_change_password,
        },
    })


@app.route("/api/logout", methods=["POST"])
def logout():
    """用户登出, 注销 token"""
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        token = auth[7:]
        TOKENS.pop(token, None)
    return jsonify({"success": True})


@app.route("/api/me", methods=["GET"])
def get_me():
    """获取当前登录用户信息（含 id 与改密标志，用于前端")"""
    user = get_current_user()
    if not user:
        return jsonify({"success": False, "error": "未登录"}), 401
    result = {
        "username": user["username"],
        "role": user["role"],
        "display_name": user["display_name"],
        "id": user.get("user_id"),
    }
    # 补充改密标志
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT must_change_password FROM users WHERE id = ?", (user.get("user_id"),))
        row = cur.fetchone()
        if row:
            result["must_change_password"] = bool(row["must_change_password"])
    except sqlite3.Error:
        pass
    finally:
        conn.close()
    return jsonify({"success": True, "user": result})


@app.route("/api/change-password", methods=["POST"])
@require_auth
def change_password():
    """
    修改当前登录用户的密码。
    请求体: {old_password: str, new_password: str}
    - 若用户为受控改密状态（must_change_password=1，如默认密码），可免填旧密码。
    - 修改成功后清除改密标志。
    """
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({"success": False, "error": "请求体必须为 JSON 格式"}), 400

    new_pwd = data.get("new_password") or ""
    old_pwd = data.get("old_password") or ""

    if not new_pwd:
        return jsonify({"success": False, "error": "请输入新密码"}), 400
    if len(new_pwd) < 4:
        return jsonify({"success": False, "error": "新密码长度至少为 4 位"}), 400
    if old_pwd and old_pwd == new_pwd:
        return jsonify({"success": False, "error": "新密码不能与原密码相同"}), 400

    user = request.current_user
    user_id = user.get("user_id")
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM users WHERE id = ?", (user_id,))
    row = cursor.fetchone()
    if row is None:
        conn.close()
        return jsonify({"success": False, "error": "用户不存在"}), 404

    must_change = bool(row["must_change_password"]) if "must_change_password" in row.keys() else False
    # 非受控改密状态下必须校验原密码
    if not must_change:
        if not old_pwd or not verify_password(row["password"], old_pwd):
            conn.close()
            return jsonify({"success": False, "error": "原密码不正确"}), 400

    cursor.execute(
        "UPDATE users SET password = ?, must_change_password = 0 WHERE id = ?",
        (hash_password(new_pwd), row["id"]),
    )
    conn.commit()
    conn.close()

    return jsonify({"success": True, "message": "密码修改成功"})


# ============================================================
# API: 用户管理 (仅管理员)
# ============================================================

@app.route("/api/users", methods=["GET"])
@require_admin
def list_users():
    """获取所有用户列表"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, username, display_name, role, created_at FROM users ORDER BY id"
    )
    rows = cursor.fetchall()
    conn.close()
    users = [row_to_dict(r) for r in rows]
    return jsonify({"success": True, "count": len(users), "data": users})


@app.route("/api/users", methods=["POST"])
@require_admin
def create_user():
    """新增用户"""
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({"success": False, "error": "请求体必须为 JSON 格式"}), 400

    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    display_name = (data.get("display_name") or "").strip() or None
    role = (data.get("role") or "user").strip().lower()

    if not username or not password:
        return jsonify({"success": False, "error": "用户名和密码不能为空"}), 400
    if role not in ("admin", "user"):
        return jsonify({"success": False, "error": "角色只能为 admin 或 user"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO users (username, password, display_name, role) VALUES (?, ?, ?, ?)",
            (username, hash_password(password), display_name, role),
        )
        conn.commit()
        new_id = cursor.lastrowid
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"success": False, "error": f"用户名 '{username}' 已存在"}), 400
    conn.close()

    return jsonify({
        "success": True,
        "message": f"成功创建用户 '{username}'",
        "id": new_id,
        "user": {"username": username, "display_name": display_name, "role": role},
    }), 201


@app.route("/api/users/<int:user_id>", methods=["PUT"])
@require_admin
def update_user(user_id):
    """修改用户 (可改密码、display_name、role)"""
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({"success": False, "error": "请求体必须为 JSON 格式"}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM users WHERE id = ?", (user_id,))
    user = cursor.fetchone()
    if user is None:
        conn.close()
        return jsonify({"success": False, "error": f"未找到 id={user_id} 的用户"}), 404

    update_fields = {}
    if "password" in data and data["password"]:
        update_fields["password"] = hash_password(data["password"])
    if "display_name" in data:
        dn = (data.get("display_name") or "").strip()
        update_fields["display_name"] = dn if dn else None
    if "role" in data:
        role = (data.get("role") or "").strip().lower()
        if role not in ("admin", "user"):
            conn.close()
            return jsonify({"success": False, "error": "角色只能为 admin 或 user"}), 400
        update_fields["role"] = role

    if not update_fields:
        conn.close()
        return jsonify({"success": False, "error": "没有可更新的字段"}), 400

    set_clause = ", ".join(f'"{k}" = ?' for k in update_fields.keys())
    values = list(update_fields.values()) + [user_id]
    cursor.execute(f'UPDATE users SET {set_clause} WHERE id = ?', values)
    conn.commit()
    conn.close()

    return jsonify({
        "success": True,
        "message": f"成功更新用户 id={user_id}",
        "id": user_id,
        "updated_fields": list(update_fields.keys()),
    })


@app.route("/api/users/<int:user_id>", methods=["DELETE"])
@require_admin
def delete_user(user_id):
    """删除用户 (不能删除自己)"""
    current = request.current_user
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT username FROM users WHERE id = ?", (user_id,))
    user = cursor.fetchone()
    if user is None:
        conn.close()
        return jsonify({"success": False, "error": f"未找到 id={user_id} 的用户"}), 404

    if user["username"] == current["username"]:
        conn.close()
        return jsonify({"success": False, "error": "不能删除当前登录的自己"}), 400

    cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()

    return jsonify({
        "success": True,
        "message": f"成功删除用户 '{user['username']}'",
        "id": user_id,
    })


# ============================================================
# API: 获取所有表信息
# ============================================================

@app.route("/api/tables", methods=["GET"])
def get_tables():
    """获取所有表名和字段定义（仅返回用户可见表，不含内部表）"""
    result = []
    for name, info in get_user_tables().items():
        result.append({
            "table_name": name,
            "name_cn": info["name_cn"],
            "field_count": len(info["fields"]),
            "fields": info["fields"],
        })
    return jsonify({
        "success": True,
        "count": len(result),
        "tables": result,
    })


# ============================================================
# API: 跨表公式计算（eval_rooms）
# ============================================================

@app.route("/api/eval_rooms/calc", methods=["POST"])
@require_auth
def calc_eval_rooms_api():
    """计算 eval_rooms 表的跨表公式值"""
    data = request.get_json(silent=True) or {}
    stat_scope = str(data.get('stat_scope', '') or '').strip()

    if not stat_scope:
        return jsonify({
            "success": True,
            "total_eval_count": '',
            "es_booking_count": '',
            "mobile_eval_count": '',
        })

    conn = get_db()
    row_data = {
        'stat_scope': stat_scope,
        'g_mobile_count': data.get('g_mobile_count', 0),
        'resource_lack_count': data.get('resource_lack_count', 0),
        'other_reason_count': data.get('other_reason_count', 0),
    }
    result = compute_formula_fields('eval_rooms', row_data, db_conn=conn)
    conn.close()

    return jsonify({
        "success": True,
        "total_eval_count": result.get('total_eval_count', ''),
        "es_booking_count": result.get('es_booking_count', ''),
        "mobile_eval_count": result.get('mobile_eval_count', ''),
        "should_use_mobile_count": result.get('should_use_mobile_count', ''),
    })


# ============================================================
# API: 获取指定表的字段定义和下拉菜单值
# ============================================================

@app.route("/api/columns/<table_name>", methods=["GET"])
@table_exists_decorator
def get_columns(table_name):
    """获取字段定义和下拉菜单值（合并 schema 字段和自定义列）"""
    # 1. 获取 schema 可见字段（排除 hidden 字段如 id）
    fields = get_visible_fields_with_custom(table_name)
    dropdowns = get_dropdown_fields(table_name)
    pk = get_primary_key(table_name)

    conn = get_db()
    cursor = conn.cursor()

    # 2. 查询该表的自定义列
    cursor.execute(
        "SELECT * FROM _custom_columns WHERE table_name = ? ORDER BY sort_order, id",
        (table_name,),
    )
    custom_rows = cursor.fetchall()

    # 3. 从数据库查询该表的自定义下拉选项, 并与 schema 默认选项合并去重
    cursor.execute(
        "SELECT field_key, option_value FROM dropdown_options "
        "WHERE table_name = ? ORDER BY field_key, sort_order, id",
        (table_name,),
    )
    db_options = {}
    for row in cursor.fetchall():
        fk = row["field_key"]
        ov = row["option_value"]
        db_options.setdefault(fk, []).append(ov)

    conn.close()

    # 4. 合并: schema 默认 options + 数据库自定义 options, 去重
    for fk, db_opts in db_options.items():
        if fk in dropdowns:
            merged = list(dropdowns[fk]["options"])
        else:
            # 数据库中存在但 schema 中没有 options 定义的字段
            field_def = get_field_by_key(table_name, fk)
            # 也检查是否为自定义列
            custom_def = None
            for cr in custom_rows:
                if cr["field_key"] == fk:
                    custom_def = cr
                    break

            if field_def:
                dropdowns[fk] = {
                    "name_cn": field_def["name_cn"],
                    "options": [],
                    "is_custom_dropdown": True,
                }
            elif custom_def:
                dropdowns[fk] = {
                    "name_cn": custom_def["field_name_cn"],
                    "options": [],
                    "is_custom_dropdown": True,
                }
            else:
                dropdowns[fk] = {
                    "name_cn": fk,
                    "options": [],
                    "is_custom_dropdown": True,
                }
            merged = []
        for opt in db_opts:
            if opt not in merged:
                merged.append(opt)
        dropdowns[fk]["options"] = merged

    # 5. 为自定义列中 is_dropdown="是" 的字段添加到 dropdowns（即使还没有选项）
    for row in custom_rows:
        fk = row["field_key"]
        if row["is_dropdown"] == "是" and fk not in dropdowns:
            dropdowns[fk] = {
                "name_cn": row["field_name_cn"],
                "options": [],
                "is_custom_dropdown": True,
            }

    # 6. 同步更新 fields 中对应字段的 options
    for field in fields:
        if field["key"] in dropdowns:
            field["options"] = dropdowns[field["key"]]["options"]

    # 构建 field_keys 和 field_cn_names（仅可见字段 + 自定义列）
    field_keys = [f["key"] for f in fields]
    field_cn_names = [f["name_cn"] for f in fields]

    return jsonify({
        "success": True,
        "table_name": table_name,
        "name_cn": get_table_cn_name(table_name),
        "primary_key": pk,
        "fields": fields,
        "dropdowns": dropdowns,
        "field_keys": field_keys,
        "field_cn_names": field_cn_names,
    })


# ============================================================
# API: 基础列隐藏/恢复 + 全部列查询
# ============================================================

# 系统列（不允许隐藏的字段 key）
SYSTEM_FIELD_KEYS = {"id"}


@app.route("/api/columns/hide", methods=["POST"])
@require_admin
def hide_base_column():
    """隐藏基础列（逻辑隐藏，不从数据库表中删除列）
    请求体: {table_name, field_key}
    """
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({"success": False, "error": "请求体必须为 JSON 格式"}), 400

    table_name = (data.get("table_name") or "").strip()
    field_key = (data.get("field_key") or "").strip()

    if not table_name or not field_key:
        return jsonify({"success": False, "error": "table_name 和 field_key 不能为空"}), 400

    if table_name not in TABLES or is_internal_table(table_name):
        return jsonify({"success": False, "error": f"无效的表名: {table_name}"}), 400

    # 检查 field_key 是否为该表的基础列（schema 中定义的列）
    field_def = get_field_by_key(table_name, field_key)
    if field_def is None:
        return jsonify({"success": False, "error": f"字段 '{field_key}' 不是表 '{table_name}' 的基础列"}), 400

    # 系统列（如 id）不允许隐藏
    pk = get_primary_key(table_name)
    if field_key in SYSTEM_FIELD_KEYS or field_key == pk:
        return jsonify({"success": False, "error": f"系统列 '{field_key}' 不允许隐藏"}), 400

    # 已经是 hidden 的字段不需要再隐藏
    if field_def.get("hidden"):
        return jsonify({"success": False, "error": f"字段 '{field_key}' 已经是隐藏状态"}), 400

    conn = get_db()
    cursor = conn.cursor()

    # 检查是否已隐藏
    cursor.execute(
        "SELECT id FROM _hidden_columns WHERE table_name = ? AND field_key = ?",
        (table_name, field_key),
    )
    existing = cursor.fetchone()
    if existing:
        conn.close()
        return jsonify({"success": False, "error": f"基础列 '{field_key}' 已处于隐藏状态"}), 400

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cursor.execute(
        "INSERT INTO _hidden_columns (table_name, field_key, hidden_at) VALUES (?, ?, ?)",
        (table_name, field_key, now),
    )
    conn.commit()
    conn.close()

    return jsonify({
        "success": True,
        "message": f"成功隐藏基础列 '{field_def['name_cn']}'",
        "table_name": table_name,
        "field_key": field_key,
        "field_name_cn": field_def["name_cn"],
    })


@app.route("/api/columns/restore", methods=["POST"])
@require_admin
def restore_base_column():
    """恢复已隐藏的基础列
    请求体: {table_name, field_key}
    """
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({"success": False, "error": "请求体必须为 JSON 格式"}), 400

    table_name = (data.get("table_name") or "").strip()
    field_key = (data.get("field_key") or "").strip()

    if not table_name or not field_key:
        return jsonify({"success": False, "error": "table_name 和 field_key 不能为空"}), 400

    if table_name not in TABLES or is_internal_table(table_name):
        return jsonify({"success": False, "error": f"无效的表名: {table_name}"}), 400

    conn = get_db()
    cursor = conn.cursor()

    # 检查是否存在隐藏记录
    cursor.execute(
        "SELECT id FROM _hidden_columns WHERE table_name = ? AND field_key = ?",
        (table_name, field_key),
    )
    row = cursor.fetchone()
    if row is None:
        conn.close()
        return jsonify({"success": False, "error": f"基础列 '{field_key}' 未处于隐藏状态"}), 400

    cursor.execute(
        "DELETE FROM _hidden_columns WHERE table_name = ? AND field_key = ?",
        (table_name, field_key),
    )
    conn.commit()
    conn.close()

    field_def = get_field_by_key(table_name, field_key)
    field_name_cn = field_def["name_cn"] if field_def else field_key

    return jsonify({
        "success": True,
        "message": f"成功恢复基础列 '{field_name_cn}'",
        "table_name": table_name,
        "field_key": field_key,
        "field_name_cn": field_name_cn,
    })


@app.route("/api/columns/rename", methods=["POST"])
@require_auth
@require_admin
def rename_column():
    """重命名表头列（管理员）
    接收: {table_name, field_key, custom_name}
    """
    data = request.get_json(force=True, silent=True) or {}
    table_name = (data.get("table_name") or "").strip()
    field_key = (data.get("field_key") or "").strip()
    custom_name = (data.get("custom_name") or "").strip()

    if not table_name or not field_key or not custom_name:
        return jsonify({"success": False, "error": "table_name, field_key, custom_name 不能为空"}), 400

    if table_name not in TABLES or is_internal_table(table_name):
        return jsonify({"success": False, "error": f"表 '{table_name}' 不存在"}), 404

    # 获取原始名称
    original_name = field_key
    field_def = get_field_by_key(table_name, field_key)
    if field_def:
        original_name = field_def.get("name_cn", field_key)
    else:
        # 可能是自定义列
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT field_name_cn FROM _custom_columns WHERE table_name = ? AND field_key = ?",
            (table_name, field_key),
        )
        row = cursor.fetchone()
        conn.close()
        if row:
            original_name = row["field_name_cn"]
        else:
            return jsonify({"success": False, "error": f"字段 '{field_key}' 不存在"}), 404

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO _column_renames (table_name, field_key, custom_name, original_name) "
        "VALUES (?, ?, ?, ?) "
        "ON CONFLICT(table_name, field_key) DO UPDATE SET custom_name = ?, original_name = ?, updated_at = datetime('now', 'localtime')",
        (table_name, field_key, custom_name, original_name, custom_name, original_name),
    )
    conn.commit()
    conn.close()

    return jsonify({
        "success": True,
        "message": f"已将 '{original_name}' 重命名为 '{custom_name}'",
        "table_name": table_name,
        "field_key": field_key,
        "custom_name": custom_name,
        "original_name": original_name,
    })


@app.route("/api/columns/rename", methods=["DELETE"])
@require_auth
@require_admin
def reset_column_name():
    """恢复表头列原始名称（管理员）
    接收: {table_name, field_key}
    """
    data = request.get_json(force=True, silent=True) or {}
    table_name = (data.get("table_name") or "").strip()
    field_key = (data.get("field_key") or "").strip()

    if not table_name or not field_key:
        return jsonify({"success": False, "error": "table_name, field_key 不能为空"}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT original_name FROM _column_renames WHERE table_name = ? AND field_key = ?",
        (table_name, field_key),
    )
    row = cursor.fetchone()
    if row is None:
        conn.close()
        return jsonify({"success": False, "error": f"字段 '{field_key}' 未被重命名"}), 400

    original_name = row["original_name"]
    cursor.execute(
        "DELETE FROM _column_renames WHERE table_name = ? AND field_key = ?",
        (table_name, field_key),
    )
    conn.commit()
    conn.close()

    return jsonify({
        "success": True,
        "message": f"已恢复原始名称 '{original_name}'",
        "table_name": table_name,
        "field_key": field_key,
        "original_name": original_name,
    })


@app.route("/api/columns/all/<table_name>", methods=["GET"])
@require_auth
@table_exists_decorator
def get_all_columns(table_name):
    """获取该表所有列的完整信息，包括基础列和自定义列，按 sort_order 排序。
    - 基础列：schema 定义的，标记 is_base: true, is_hidden: true/false
    - 自定义列：从 _custom_columns 查的，标记 is_base: false, is_custom: true
    """
    # 获取隐藏的基础列 key 集合
    hidden_keys = get_hidden_field_keys(table_name)

    # 获取所有 schema 字段（包括 hidden 的如 id）
    all_base_fields = get_fields(table_name)
    pk = get_primary_key(table_name)

    # 获取自定义列
    custom_fields = []
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT * FROM _custom_columns WHERE table_name = ? ORDER BY sort_order, id",
            (table_name,),
        )
        for row in cursor.fetchall():
            after_field = row["after_field"] if "after_field" in row.keys() else None
            custom_fields.append({
                "key": row["field_key"],
                "name_cn": row["field_name_cn"],
                "type": row["field_type"],
                "required": False,
                "is_base": False,
                "is_custom": True,
                "is_hidden": False,
                "after_field": after_field,
                "sort_order": row["sort_order"],
                "custom_id": row["id"],
            })
        conn.close()
    except sqlite3.Error:
        pass

    # 构建完整列表：基础列（含 hidden 的）+ 自定义列，按 sort_order 排序
    result = []
    placed_custom_keys = set()

    # 第一轮：遍历所有基础列，插入 after_field 指向当前基础列的自定义列
    for idx, bf in enumerate(all_base_fields):
        field_info = dict(bf)
        field_info["is_base"] = True
        field_info["is_custom"] = False
        field_info["is_hidden"] = bf.get("hidden", False) or (bf["key"] in hidden_keys)
        field_info["is_system"] = bf["key"] in SYSTEM_FIELD_KEYS or bf["key"] == pk
        field_info["sort_order"] = idx
        result.append(field_info)

        # 如果该基础列未被隐藏，查找 after_field 指向它的自定义列
        customs_after = [
            cf for cf in custom_fields
            if cf.get("after_field") == bf["key"] and cf["key"] not in placed_custom_keys
        ]
        for cf in customs_after:
            result.append(cf)
            placed_custom_keys.add(cf["key"])

    # 第二轮：处理 after_field 指向其他自定义列的情况（链式插入）
    changed = True
    max_iterations = len(custom_fields) + 1
    iteration = 0
    while changed and iteration < max_iterations:
        changed = False
        iteration += 1
        for cf in custom_fields:
            if cf["key"] in placed_custom_keys:
                continue
            after = cf.get("after_field")
            if after:
                for i, item in enumerate(result):
                    if item["key"] == after:
                        result.insert(i + 1, cf)
                        placed_custom_keys.add(cf["key"])
                        changed = True
                        break

    # 第三轮：将未放置的自定义列追加到末尾
    for cf in custom_fields:
        if cf["key"] not in placed_custom_keys:
            result.append(cf)
            placed_custom_keys.add(cf["key"])

    # 应用自定义重命名
    apply_column_renames(table_name, result)

    # 获取重命名信息，为每列添加 original_name_cn 和 is_renamed
    renames = get_column_renames(table_name)
    for col in result:
        key = col.get("key", "")
        if key in renames:
            if not col.get("original_name_cn"):
                col["original_name_cn"] = col.get("name_cn", "")
            col["name_cn"] = renames[key]
            col["is_renamed"] = True
        else:
            col["is_renamed"] = False

    return jsonify({
        "success": True,
        "table_name": table_name,
        "name_cn": get_table_cn_name(table_name),
        "primary_key": pk,
        "columns": result,
        "count": len(result),
        "hidden_count": sum(1 for c in result if c.get("is_hidden")),
        "custom_count": len(custom_fields),
        "renamed_count": len(renames),
    })


# ============================================================
# API: 下拉选项管理
# ============================================================

@app.route("/api/dropdowns", methods=["GET"])
@require_auth
def get_all_dropdowns():
    """获取所有下拉选项, 返回 {table_name: {field_key: [options]}} 格式"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, table_name, field_key, option_value FROM dropdown_options "
        "ORDER BY table_name, field_key, sort_order, id"
    )
    result = {}
    for row in cursor.fetchall():
        tn = row["table_name"]
        fk = row["field_key"]
        ov = row["option_value"]
        result.setdefault(tn, {}).setdefault(fk, []).append({"id": row["id"], "value": ov})
    conn.close()
    return jsonify({"success": True, "data": result})


@app.route("/api/dropdowns/<table_name>", methods=["GET"])
@require_auth
def get_table_dropdowns(table_name):
    """获取指定表的下拉选项"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, field_key, option_value FROM dropdown_options "
        "WHERE table_name = ? ORDER BY field_key, sort_order, id",
        (table_name,),
    )
    result = {}
    for row in cursor.fetchall():
        fk = row["field_key"]
        ov = row["option_value"]
        result.setdefault(fk, []).append({"id": row["id"], "value": ov})
    conn.close()
    return jsonify({"success": True, "table_name": table_name, "data": result})


@app.route("/api/dropdowns", methods=["POST"])
@require_admin
def add_dropdown_option():
    """添加下拉选项"""
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({"success": False, "error": "请求体必须为 JSON 格式"}), 400

    table_name = (data.get("table_name") or "").strip()
    field_key = (data.get("field_key") or "").strip()
    option_value = (data.get("option_value") or "").strip()

    if not table_name or not field_key or not option_value:
        return jsonify({"success": False, "error": "table_name, field_key, option_value 不能为空"}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO dropdown_options (table_name, field_key, option_value) VALUES (?, ?, ?)",
            (table_name, field_key, option_value),
        )
        conn.commit()
        new_id = cursor.lastrowid
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"success": False, "error": "该选项已存在"}), 400
    conn.close()

    return jsonify({
        "success": True,
        "message": f"成功添加选项 '{option_value}'",
        "id": new_id,
        "option": {"table_name": table_name, "field_key": field_key, "option_value": option_value},
    }), 201


@app.route("/api/dropdowns/<int:option_id>", methods=["DELETE"])
@require_admin
def delete_dropdown_option(option_id):
    """删除下拉选项"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM dropdown_options WHERE id = ?", (option_id,))
    row = cursor.fetchone()
    if row is None:
        conn.close()
        return jsonify({"success": False, "error": f"未找到 id={option_id} 的选项"}), 404

    cursor.execute("DELETE FROM dropdown_options WHERE id = ?", (option_id,))
    conn.commit()
    conn.close()

    return jsonify({
        "success": True,
        "message": f"成功删除选项 id={option_id}",
        "id": option_id,
    })


@app.route("/api/dropdowns/enable", methods=["POST"])
@require_admin
def enable_dropdown_field():
    """将非下拉字段转为下拉字段（为任意字段启用下拉功能）"""
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({"success": False, "error": "请求体必须为 JSON 格式"}), 400

    table_name = (data.get("table_name") or "").strip()
    field_key = (data.get("field_key") or "").strip()

    if not table_name or not field_key:
        return jsonify({"success": False, "error": "table_name 和 field_key 不能为空"}), 400

    if table_name not in TABLES or is_internal_table(table_name):
        return jsonify({"success": False, "error": "无效的表名"}), 400

    conn = get_db()
    cursor = conn.cursor()

    # 检查该字段是否为自定义列
    cursor.execute(
        "SELECT * FROM _custom_columns WHERE table_name = ? AND field_key = ?",
        (table_name, field_key),
    )
    custom_col = cursor.fetchone()
    is_custom = False

    if custom_col:
        is_custom = True
        # 更新自定义列的 is_dropdown 标记
        cursor.execute(
            "UPDATE _custom_columns SET is_dropdown = '是' WHERE table_name = ? AND field_key = ?",
            (table_name, field_key),
        )

    # 检查 dropdown_options 表中是否已有该字段的记录
    cursor.execute(
        "SELECT COUNT(*) FROM dropdown_options WHERE table_name = ? AND field_key = ?",
        (table_name, field_key),
    )
    option_count = cursor.fetchone()[0]

    conn.commit()
    conn.close()

    return jsonify({
        "success": True,
        "message": f"字段 '{field_key}' 已启用下拉功能",
        "is_custom_column": is_custom,
        "has_options": option_count > 0,
    })


# ============================================================
# API: 自定义列管理 (R8)
# ============================================================

@app.route("/api/custom-columns", methods=["GET"])
@app.route("/api/custom-columns/<table_name>", methods=["GET"])
@require_auth
def get_custom_columns(table_name=None):
    """获取自定义列，可按表名过滤"""
    conn = get_db()
    cursor = conn.cursor()
    if table_name:
        cursor.execute(
            "SELECT id, table_name, field_key, field_name_cn, field_type, is_dropdown, sort_order, after_field, created_at "
            "FROM _custom_columns WHERE table_name = ? ORDER BY sort_order, id",
            (table_name,)
        )
    else:
        cursor.execute(
            "SELECT id, table_name, field_key, field_name_cn, field_type, is_dropdown, sort_order, after_field, created_at "
            "FROM _custom_columns ORDER BY table_name, sort_order, id"
        )
    rows = cursor.fetchall()
    conn.close()

    if table_name:
        # 单表模式: 返回数组格式，兼容前端多种解析方式
        columns = [{
            "id": row["id"],
            "table_name": row["table_name"],
            "field_key": row["field_key"],
            "field_name_cn": row["field_name_cn"],
            "field_type": row["field_type"],
            "is_dropdown": row["is_dropdown"],
            "sort_order": row["sort_order"],
            "after_field": row["after_field"] if "after_field" in row.keys() else None,
            "created_at": row["created_at"],
        } for row in rows]
        return jsonify({"success": True, "count": len(columns), "columns": columns, "data": columns})

    # 全部模式: 按表分组
    result = {}
    for row in rows:
        tn = row["table_name"]
        result.setdefault(tn, []).append({
            "id": row["id"],
            "table_name": tn,
            "field_key": row["field_key"],
            "field_name_cn": row["field_name_cn"],
            "field_type": row["field_type"],
            "is_dropdown": row["is_dropdown"],
            "sort_order": row["sort_order"],
            "after_field": row["after_field"] if "after_field" in row.keys() else None,
            "created_at": row["created_at"],
        })

    return jsonify({"success": True, "count": len(rows), "data": result})


@app.route("/api/custom-columns", methods=["POST"])
@require_admin
def add_custom_column():
    """添加自定义列
    可选参数 after_field: 表示在哪个字段后面插入，影响 sort_order
    """
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({"success": False, "error": "请求体必须为 JSON 格式"}), 400

    table_name = (data.get("table_name") or "").strip()
    field_name_cn = (data.get("field_name_cn") or data.get("column_name") or "").strip()
    field_type = (data.get("field_type") or data.get("column_type") or "string").strip()
    is_dropdown = (data.get("is_dropdown") or "否").strip()
    after_field = (data.get("after_field") or "").strip() or None

    if not table_name or not field_name_cn:
        return jsonify({"success": False, "error": "table_name 和 field_name_cn 不能为空"}), 400

    if table_name not in TABLES or is_internal_table(table_name):
        return jsonify({"success": False, "error": f"无效的表名: {table_name}"}), 400

    if field_type not in ("string", "number", "integer", "date"):
        return jsonify({"success": False, "error": f"无效的字段类型: {field_type}"}), 400

    conn = get_db()
    cursor = conn.cursor()

    # 生成 field_key（使用 col_1, col_2 等格式，避免与 schema 字段冲突）
    existing_keys = get_field_keys(table_name)
    cursor.execute("SELECT COUNT(*) FROM _custom_columns WHERE table_name = ?", (table_name,))
    count = cursor.fetchone()[0]
    field_key = f"col_{count + 1}"
    while field_key in existing_keys:
        count += 1
        field_key = f"col_{count + 1}"

    # 检查实际表中是否已有该列
    cursor.execute(f'PRAGMA table_info("{table_name}")')
    actual_columns = [row[1] for row in cursor.fetchall()]
    if field_key in actual_columns:
        conn.close()
        return jsonify({"success": False, "error": f"字段 '{field_key}' 已存在于表中"}), 400

    # 校验 after_field 是否为该表的有效字段（基础列或已存在的自定义列）
    if after_field:
        all_valid_keys = get_all_field_keys_with_custom(table_name)
        if after_field not in all_valid_keys:
            conn.close()
            return jsonify({"success": False, "error": f"after_field '{after_field}' 不是表 '{table_name}' 中的有效字段"}), 400

    # 计算 sort_order
    if after_field:
        # 查找 after_field 的 sort_order
        after_sort_order = None
        # 先检查是否为基础列
        base_fields = get_visible_fields(table_name)
        for idx, f in enumerate(base_fields):
            if f["key"] == after_field:
                after_sort_order = idx
                break
        if after_sort_order is None:
            # 检查是否为自定义列
            cursor.execute(
                "SELECT sort_order FROM _custom_columns WHERE table_name = ? AND field_key = ?",
                (table_name, after_field),
            )
            row = cursor.fetchone()
            if row:
                after_sort_order = row["sort_order"]

        if after_sort_order is not None:
            # 在同一 after_field 后已有多少自定义列
            cursor.execute(
                "SELECT COUNT(*) FROM _custom_columns WHERE table_name = ? AND after_field = ?",
                (table_name, after_field),
            )
            after_count = cursor.fetchone()[0]
            sort_order = (after_sort_order + 1) * 1000 + after_count
        else:
            sort_order = count + 1
    else:
        sort_order = count + 1

    # 插入 _custom_columns 记录
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cursor.execute(
        "INSERT INTO _custom_columns (table_name, field_key, field_name_cn, field_type, is_dropdown, sort_order, after_field, created_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (table_name, field_key, field_name_cn, field_type, is_dropdown, sort_order, after_field, now),
    )
    new_id = cursor.lastrowid

    # 执行 ALTER TABLE 添加实际列
    sqlite_type = get_sqlite_type(field_type)
    try:
        cursor.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "{field_key}" {sqlite_type}')
    except sqlite3.OperationalError as e:
        conn.close()
        return jsonify({"success": False, "error": f"添加列失败: {str(e)}"}), 500

    conn.commit()
    conn.close()

    return jsonify({
        "success": True,
        "message": f"成功添加自定义列 '{field_name_cn}'",
        "id": new_id,
        "column": {
            "table_name": table_name,
            "field_key": field_key,
            "field_name_cn": field_name_cn,
            "field_type": field_type,
            "is_dropdown": is_dropdown,
            "after_field": after_field,
            "sort_order": sort_order,
        },
    }), 201


@app.route("/api/custom-columns/<int:column_id>", methods=["DELETE"])
@require_admin
def delete_custom_column(column_id):
    """删除自定义列（SQLite 不支持 DROP COLUMN，需重建表）"""
    conn = get_db()
    cursor = conn.cursor()

    # 查询列信息
    cursor.execute("SELECT * FROM _custom_columns WHERE id = ?", (column_id,))
    row = cursor.fetchone()
    if row is None:
        conn.close()
        return jsonify({"success": False, "error": f"未找到 id={column_id} 的自定义列"}), 404

    table_name = row["table_name"]
    field_key = row["field_key"]
    field_name_cn = row["field_name_cn"]

    # 获取当前表结构
    cursor.execute(f'PRAGMA table_info("{table_name}")')
    columns_info = cursor.fetchall()
    column_names = [col[1] for col in columns_info]

    if field_key in column_names:
        # 需要重建表（SQLite 不支持 DROP COLUMN）
        keep_columns = [col for col in columns_info if col[1] != field_key]

        # 构建新表列定义
        col_defs = []
        for col in keep_columns:
            col_name = col[1]
            col_type = col[2]
            col_def = f'"{col_name}" {col_type}'
            if col[5]:  # primary key
                col_def += " PRIMARY KEY"
            col_defs.append(col_def)

        temp_table = f"_temp_{table_name}"

        # 创建临时表
        cursor.execute(f'CREATE TABLE "{temp_table}" ({", ".join(col_defs)})')

        # 复制数据（排除被删列）
        keep_col_names = [col[1] for col in keep_columns]
        keep_cols_quoted = ", ".join(f'"{c}"' for c in keep_col_names)
        cursor.execute(
            f'INSERT INTO "{temp_table}" ({keep_cols_quoted}) '
            f'SELECT {keep_cols_quoted} FROM "{table_name}"'
        )

        # 删除旧表
        cursor.execute(f'DROP TABLE "{table_name}"')

        # 重命名临时表
        cursor.execute(f'ALTER TABLE "{temp_table}" RENAME TO "{table_name}"')

    # 从 _custom_columns 表删除记录
    cursor.execute("DELETE FROM _custom_columns WHERE id = ?", (column_id,))

    # 同时删除该字段的下拉选项
    cursor.execute(
        "DELETE FROM dropdown_options WHERE table_name = ? AND field_key = ?",
        (table_name, field_key),
    )

    conn.commit()
    conn.close()

    return jsonify({
        "success": True,
        "message": f"成功删除自定义列 '{field_name_cn}'",
        "id": column_id,
    })


# ============================================================
# API: CRUD 操作
# ============================================================

@app.route("/api/<table_name>", methods=["GET"])
@require_auth
@table_exists_decorator
def list_records(table_name):
    """
    获取记录列表, 支持分页和搜索。
    Query params:
      - page: 页码 (默认 1)
      - per_page: 每页条数 (默认 20)
      - search: 搜索关键词 (在所有文本字段中模糊匹配)
      - order_by: 排序字段 (默认主键)
      - order_dir: 排序方向 asc/desc (默认 asc)
    """
    page = max(1, request.args.get("page", 1, type=int))
    per_page = max(1, min(100000, request.args.get("per_page", 20, type=int)))
    search = request.args.get("search", "").strip()
    order_by = request.args.get("order_by", "").strip()
    order_dir = request.args.get("order_dir", "asc").strip().lower()

    fields = get_fields(table_name)
    field_keys = get_field_keys(table_name)
    pk = get_primary_key(table_name)

    # 构建排序
    if order_by and order_by in field_keys:
        order_field = order_by
    else:
        order_field = pk
    if order_dir not in ("asc", "desc"):
        order_dir = "asc"

    # 构建 WHERE 条件
    where_clause = ""
    params = []
    if search:
        search_conditions = []
        for f in fields:
            if f["type"] == "string" or f["type"] == "date":
                search_conditions.append(f'"{f["key"]}" LIKE ?')
                params.append(f"%{search}%")
        if search_conditions:
            where_clause = " WHERE " + " OR ".join(search_conditions)

    # 获取总数
    conn = get_db()
    cursor = conn.cursor()
    count_sql = f'SELECT COUNT(*) FROM "{table_name}"{where_clause}'
    cursor.execute(count_sql, params)
    total = cursor.fetchone()[0]

    # 分页查询
    offset = (page - 1) * per_page
    data_sql = (
        f'SELECT * FROM "{table_name}"{where_clause} '
        f'ORDER BY "{order_field}" {order_dir.upper()} '
        f'LIMIT ? OFFSET ?'
    )
    cursor.execute(data_sql, params + [per_page, offset])
    rows = cursor.fetchall()
    records = [row_to_dict(r) for r in rows]

    # 序列化
    for record in records:
        for k, v in record.items():
            record[k] = serialize_value(v)

    conn.close()

    total_pages = (total + per_page - 1) // per_page if total > 0 else 0

    return jsonify({
        "success": True,
        "table_name": table_name,
        "table_cn_name": get_table_cn_name(table_name),
        "page": page,
        "per_page": per_page,
        "total": total,
        "total_pages": total_pages,
        "data": records,
    })


@app.route("/api/<table_name>/<int:record_id>", methods=["GET"])
@require_auth
@table_exists_decorator
def get_record(table_name, record_id):
    """获取单条记录"""
    pk = get_primary_key(table_name)
    conn = get_db()
    cursor = conn.cursor()
    sql = f'SELECT * FROM "{table_name}" WHERE "{pk}" = ?'
    cursor.execute(sql, (record_id,))
    row = cursor.fetchone()
    conn.close()

    if row is None:
        return jsonify({
            "success": False,
            "error": f"未找到 {table_name} 中 {pk}={record_id} 的记录",
        }), 404

    record = row_to_dict(row)
    for k, v in record.items():
        record[k] = serialize_value(v)

    return jsonify({
        "success": True,
        "data": record,
    })


@app.route("/api/<table_name>", methods=["POST"])
@require_auth
@table_exists_decorator
def create_record(table_name):
    """新增记录"""
    data = request.get_json(silent=True)
    if data is None:
        return jsonify({"success": False, "error": "请求体必须为 JSON 格式"}), 400

    if not isinstance(data, dict):
        return jsonify({"success": False, "error": "请求体必须为 JSON 对象"}), 400

    fields = get_fields(table_name)
    field_keys = get_all_field_keys_with_custom(table_name)
    pk = get_primary_key(table_name)

    # 清理数据
    record = clean_record(data)

    # 过滤掉不在字段列表中的 key。
    # 主键交由 SQLite 自增（rowid/AUTOINCREMENT）分配，避免并发下 MAX(pk)+1 竞争冲突；
    # 若调用方显式提供了主键，将随 filtered 一并插入。
    filtered = {k: v for k, v in record.items() if k in field_keys}

    # 自动记录创建者
    filtered["created_by"] = request.current_user["username"]
    # 自动记录创建时间（即首次修改时间）—— 北京时间
    filtered["updated_at"] = now_beijing()

    # 自动计算公式列初始值（如果用户未手动填写）
    # eval_rooms 需要DB连接进行跨表查询，先获取连接
    conn = get_db()
    formula_values = compute_formula_fields(table_name, filtered, db_conn=conn)
    for fkey, fval in formula_values.items():
        if fkey in field_keys:
            current_val = filtered.get(fkey)
            if current_val is None or current_val == '':
                filtered[fkey] = fval

    # 插入
    cursor = conn.cursor()

    # seq_no 唯一性检查 (R2): 同一账号可重复使用相同序号，不同账号不能使用相同序号
    user = request.current_user
    if user["role"] != "admin":
        seq_fields = get_seq_fields(table_name)
        if seq_fields:
            seq_field = seq_fields[0]
            seq_value = filtered.get(seq_field["key"])
            if seq_value:
                cursor.execute(
                    f'SELECT created_by FROM "{table_name}" WHERE "{seq_field["key"]}" = ? AND created_by != ?',
                    (seq_value, user["username"]),
                )
                existing = cursor.fetchone()
                if existing:
                    conn.close()
                    return jsonify({
                        "success": False,
                        "error": f"序号 '{seq_value}' 已被账号 '{existing['created_by']}' 使用，请使用其他序号",
                    }), 400

    columns = list(filtered.keys())
    placeholders = ["?"] * len(columns)
    values = [filtered[c] for c in columns]

    col_list = ", ".join('"' + c + '"' for c in columns)
    sql = (
        f'INSERT INTO "{table_name}" '
        f'({col_list}) '
        f'VALUES ({", ".join(placeholders)})'
    )
    try:
        cursor.execute(sql, values)
        conn.commit()
        new_id = filtered.get(pk, cursor.lastrowid)
        conn.close()

        # 查询返回新记录
        return jsonify({
            "success": True,
            "message": f"成功在「{get_table_cn_name(table_name)}」中新增记录",
            "id": new_id,
            "data": filtered,
        }), 201

    except sqlite3.IntegrityError as e:
        conn.close()
        return jsonify({
            "success": False,
            "error": f"数据完整性错误: {str(e)}",
        }), 400
    except sqlite3.Error as e:
        conn.close()
        return jsonify({
            "success": False,
            "error": f"数据库错误: {str(e)}",
        }), 500


@app.route("/api/<table_name>/<int:record_id>", methods=["PUT"])
@require_auth
@table_exists_decorator
def update_record(table_name, record_id):
    """更新记录"""
    data = request.get_json(silent=True)
    if data is None:
        return jsonify({"success": False, "error": "请求体必须为 JSON 格式"}), 400

    if not isinstance(data, dict):
        return jsonify({"success": False, "error": "请求体必须为 JSON 对象"}), 400

    fields = get_fields(table_name)
    field_keys = get_all_field_keys_with_custom(table_name)
    pk = get_primary_key(table_name)

    # 清理数据
    record = clean_record(data)

    # 过滤掉不在字段列表中的 key, 且不更新主键
    update_fields = {k: v for k, v in record.items() if k in field_keys and k != pk}

    if not update_fields:
        return jsonify({
            "success": False,
            "error": "没有可更新的字段",
        }), 400

    # 自动记录修改时间 —— 北京时间
    update_fields["updated_at"] = now_beijing()

    # 如果更新了公式列的输入字段，自动重新计算公式列（除非用户同时手动修改了公式列）
    conn = get_db()
    cursor = conn.cursor()

    # 记录 projects 表修改前的旧值（用于跨表重算）
    old_projects_row = None
    if table_name == 'projects':
        cursor.execute(f'SELECT * FROM "{table_name}" WHERE "{pk}" = ?', (record_id,))
        old_row_data = cursor.fetchone()
        if old_row_data:
            old_projects_row = dict(old_row_data)

    recalc_fields = should_recalc_on_update(table_name, update_fields)
    if recalc_fields:
        # 获取当前完整行数据（数据库中的旧值 + 本次更新的新值）
        cursor.execute(f'SELECT * FROM "{table_name}" WHERE "{pk}" = ?', (record_id,))
        existing = dict(cursor.fetchone() or {})
        # 合并：用新值覆盖旧值
        merged_row = {**existing, **update_fields}
        formula_values = compute_formula_fields(table_name, merged_row, db_conn=conn)
        for fkey, fval in formula_values.items():
            if fkey in recalc_fields and fkey not in update_fields:
                # 只有用户没有手动修改此公式列时才自动更新
                update_fields[fkey] = fval

    # 检查记录是否存在并获取创建者
    cursor.execute(f'SELECT created_by FROM "{table_name}" WHERE "{pk}" = ?', (record_id,))
    row = cursor.fetchone()
    if row is None:
        conn.close()
        return jsonify({
            "success": False,
            "error": f"未找到 {table_name} 中 {pk}={record_id} 的记录",
        }), 404

    # 权限检查: 管理员可编辑任何记录, 普通用户只能编辑自己创建的记录
    user = request.current_user
    created_by = row["created_by"]
    if user["role"] != "admin" and created_by != user["username"]:
        conn.close()
        return jsonify({
            "success": False,
            "error": "无权编辑其他用户创建的记录",
        }), 403

    # seq_no 唯一性检查 (R2): 同一账号可重复使用相同序号，不同账号不能使用相同序号
    if user["role"] != "admin":
        seq_fields = get_seq_fields(table_name)
        if seq_fields:
            seq_field = seq_fields[0]
            seq_value = update_fields.get(seq_field["key"])
            if seq_value:
                cursor.execute(
                    f'SELECT created_by FROM "{table_name}" '
                    f'WHERE "{seq_field["key"]}" = ? AND created_by != ? AND "{pk}" != ?',
                    (seq_value, user["username"], record_id),
                )
                existing = cursor.fetchone()
                if existing:
                    conn.close()
                    return jsonify({
                        "success": False,
                        "error": f"序号 '{seq_value}' 已被账号 '{existing['created_by']}' 使用，请使用其他序号",
                    }), 400

    # 更新
    set_clause = ", ".join(f'"{k}" = ?' for k in update_fields.keys())
    values = list(update_fields.values()) + [record_id]

    sql = f'UPDATE "{table_name}" SET {set_clause} WHERE "{pk}" = ?'
    try:
        cursor.execute(sql, values)
        conn.commit()

        # 跨表公式重算：当 projects 表记录更新后，重算依赖的 eval_rooms 记录
        if table_name == 'projects':
            # 获取更新后的新行数据
            new_projects_row = {**old_projects_row, **update_fields} if old_projects_row else update_fields
            dependent_rooms = get_dependent_eval_rooms(conn, table_name, old_row=old_projects_row, new_row=new_projects_row)
            for room_id, room_scope in dependent_rooms:
                # 获取 eval_rooms 行数据
                cursor.execute('SELECT * FROM eval_rooms WHERE id = ?', (room_id,))
                room_row = cursor.fetchone()
                if room_row:
                    room_data = dict(room_row)
                    room_formula = compute_formula_fields('eval_rooms', room_data, db_conn=conn)
                    # 更新公式列（跨表部分）
                    room_updates = {}
                    for fkey, fval in room_formula.items():
                        if fkey in ('total_eval_count', 'es_booking_count', 'mobile_eval_count'):
                            room_updates[fkey] = fval
                    if room_updates:
                        room_set = ", ".join(f'"{k}" = ?' for k in room_updates.keys())
                        room_vals = list(room_updates.values()) + [room_id]
                        cursor.execute(f'UPDATE eval_rooms SET {room_set} WHERE id = ?', room_vals)
            if dependent_rooms:
                conn.commit()

        conn.close()

        return jsonify({
            "success": True,
            "message": f"成功更新「{get_table_cn_name(table_name)}」中 {pk}={record_id} 的记录",
            "id": record_id,
            "updated_fields": list(update_fields.keys()),
        })

    except sqlite3.Error as e:
        conn.close()
        return jsonify({
            "success": False,
            "error": f"数据库错误: {str(e)}",
        }), 500


@app.route("/api/<table_name>/<int:record_id>", methods=["DELETE"])
@require_auth
@table_exists_decorator
def delete_record(table_name, record_id):
    """删除记录"""
    pk = get_primary_key(table_name)
    user = request.current_user
    conn = get_db()
    cursor = conn.cursor()

    # 检查记录是否存在并获取创建者
    cursor.execute(f'SELECT created_by FROM "{table_name}" WHERE "{pk}" = ?', (record_id,))
    row = cursor.fetchone()
    if row is None:
        conn.close()
        return jsonify({
            "success": False,
            "error": f"未找到 {table_name} 中 {pk}={record_id} 的记录",
        }), 404

    # 权限检查: 管理员可删除任何记录, 普通用户只能删除自己创建的记录
    created_by = row["created_by"]
    if user["role"] != "admin" and created_by != user["username"]:
        conn.close()
        return jsonify({
            "success": False,
            "error": "无权删除其他用户创建的记录",
        }), 403

    try:
        cursor.execute(f'DELETE FROM "{table_name}" WHERE "{pk}" = ?', (record_id,))
        conn.commit()
        conn.close()

        return jsonify({
            "success": True,
            "message": f"成功删除「{get_table_cn_name(table_name)}」中 {pk}={record_id} 的记录",
            "id": record_id,
        })

    except sqlite3.Error as e:
        conn.close()
        return jsonify({
            "success": False,
            "error": f"数据库错误: {str(e)}",
        }), 500


# ============================================================
# API: 数据校验
# ============================================================

@app.route("/api/<table_name>/validate", methods=["POST"])
@require_auth
@table_exists_decorator
def validate_records(table_name):
    """
    校验记录。
    请求体:
      - 单条: {"record": {...}}
      - 批量: {"records": [{...}, {...}]}
    """
    data = request.get_json(silent=True)
    if data is None:
        return jsonify({"success": False, "error": "请求体必须为 JSON 格式"}), 400

    if "records" in data and isinstance(data["records"], list):
        # 批量校验
        result = validate_batch(table_name, data["records"], db_path=DB_PATH)
        return jsonify({
            "success": True,
            "table_name": table_name,
            "table_cn_name": get_table_cn_name(table_name),
            "total": result["total"],
            "valid_count": result["valid_count"],
            "invalid_count": result["invalid_count"],
            "results": result["results"],
        })

    elif "record" in data and isinstance(data["record"], dict):
        # 单条校验
        result = validate_record(table_name, data["record"], db_path=DB_PATH)
        return jsonify({
            "success": True,
            "table_name": table_name,
            "table_cn_name": get_table_cn_name(table_name),
            "valid": result["valid"],
            "errors": result["errors"],
            "warnings": result["warnings"],
            "created_by": data["record"].get("created_by", "未知"),
        })

    else:
        return jsonify({
            "success": False,
            "error": "请求体必须包含 'record' 或 'records' 字段",
        }), 400


# ============================================================
# API: Excel 导出
# ============================================================

@app.route("/api/export/<table_name>", methods=["GET"])
@require_auth
@table_exists_decorator
def export_excel(table_name):
    """导出指定表的数据为 Excel 文件"""
    fields = get_visible_fields_with_custom(table_name)
    field_keys = [f["key"] for f in fields]
    field_cn_names = [f["name_cn"] for f in fields]
    pk = get_primary_key(table_name)

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(f'SELECT * FROM "{table_name}" ORDER BY "{pk}"')
    rows = cursor.fetchall()
    conn.close()

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = get_table_cn_name(table_name)

    # 表头样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 写入表头 (中文名)
    for col_idx, cn_name in enumerate(field_cn_names, 1):
        cell = ws.cell(row=1, column=col_idx, value=cn_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 最后一列: 最近修改时间
    last_col_idx = len(field_cn_names) + 1
    cell = ws.cell(row=1, column=last_col_idx, value="最近修改时间")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

    # 写入数据
    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(field_keys, 1):
            value = row[key] if key in row.keys() else None
            cell = ws.cell(row=row_idx, column=col_idx, value=serialize_value(value))
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        # 最后一列: updated_at
        updated_val = row["updated_at"] if "updated_at" in row.keys() else None
        cell = ws.cell(row=row_idx, column=last_col_idx, value=serialize_value(updated_val))
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    # 自动列宽 (限制最大宽度)
    for col_idx in range(1, last_col_idx + 1):
        cn_name = ws.cell(row=1, column=col_idx).value or ""
        max_length = len(cn_name)
        for row_idx in range(2, len(rows) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, min(len(str(cell_value)), 50))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 2

    # 冻结首行
    ws.freeze_panes = "A2"

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{table_name}_{datetime.now(BEIJING_TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/export-all", methods=["GET"])
@require_admin
def export_all_excel():
    """导出所有表的数据到一个 Excel 文件 (多 Sheet)"""
    wb = Workbook()
    # 删除默认 Sheet
    wb.remove(wb.active)

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    conn = get_db()
    cursor = conn.cursor()

    for idx, (table_name, table_info) in enumerate(get_user_tables().items(), 1):
        sheet_name = f"{idx}.{table_info['name_cn']}"
        ws = wb.create_sheet(title=sheet_name[:31])
        fields = get_visible_fields_with_custom(table_name)
        field_keys = [f["key"] for f in fields]
        field_cn_names = [f["name_cn"] for f in fields]
        pk = get_primary_key(table_name)

        # 写入表头
        for col_idx, cn_name in enumerate(field_cn_names, 1):
            cell = ws.cell(row=1, column=col_idx, value=cn_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 最后一列: 最近修改时间
        last_col_idx = len(field_cn_names) + 1
        cell = ws.cell(row=1, column=last_col_idx, value="最近修改时间")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

        # 查询数据
        cursor.execute(f'SELECT * FROM "{table_name}" ORDER BY "{pk}"')
        rows = cursor.fetchall()

        # 写入数据
        for row_idx, row in enumerate(rows, 2):
            for col_idx, key in enumerate(field_keys, 1):
                value = row[key] if key in row.keys() else None
                cell = ws.cell(row=row_idx, column=col_idx, value=serialize_value(value))
                cell.border = thin_border
            # 最后一列: updated_at
            updated_val = row["updated_at"] if "updated_at" in row.keys() else None
            cell = ws.cell(row=row_idx, column=last_col_idx, value=serialize_value(updated_val))
            cell.border = thin_border

        ws.freeze_panes = "A2"

    conn.close()

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"全表导出_{datetime.now(BEIJING_TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ============================================================
# API: Excel 导入
# ============================================================

def _coerce_value(value, field_type):
    """尝试将值转换为字段类型兼容的格式，提高导入成功率。"""
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
        if s == "":
            return None
        ft = (field_type or "").lower()
        # 数字类型: 尝试去除千分位逗号后转换
        if ft in ("number", "integer", "float", "decimal", "real"):
            cleaned = s.replace(",", "").replace("，", "")
            try:
                if ft in ("integer",):
                    return int(float(cleaned))
                return float(cleaned)
            except ValueError:
                return s  # 保留原值，SQLite 会以文本存储
        # 布尔类型
        if ft in ("boolean", "bool"):
            if s in ("是", "yes", "true", "1", "Y"):
                return "是"
            if s in ("否", "no", "false", "0", "N"):
                return "否"
        return s
    return value


def _detect_hint_row(ws):
    """检测第2行是否为格式说明行（模板生成）。返回 True 表示第2行是说明行，数据从第3行开始。"""
    if ws.max_row < 2:
        return False
    row2_values = [str(ws.cell(row=2, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    hint_keywords = ["必填", "选填", "数字", "日期", "选项"]
    hint_count = sum(1 for v in row2_values if any(kw in v for kw in hint_keywords))
    # 如果超过一半的列包含格式说明关键词，则认为是说明行
    non_empty = sum(1 for v in row2_values if v)
    if non_empty > 0 and hint_count >= non_empty * 0.5:
        return True
    return False


def _coerce_import_cell(value):
    """将 Excel 单元格值转换为便于入库的原始值（日期对象转字符串）。"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip() if isinstance(value, str) else value


def _match_sheet_to_table(sheet, table_name=None):
    """按表名参数或 Sheet 名匹配目标台账表。"""
    if table_name and table_name in TABLES and not is_internal_table(table_name):
        return table_name
    clean_sheet = sheet
    if len(sheet) > 2 and sheet[0].isdigit() and sheet[1] == ".":
        clean_sheet = sheet[2:]
    for t_name, t_info in get_user_tables().items():
        if clean_sheet == t_info["name_cn"] or sheet == t_info["name_cn"] or sheet == t_name:
            return t_name
    return None


def _import_worksheet(ws, sheet, target_table, username, do_validate, skip_errors):
    """
    将单个 Excel Sheet 数据导入目标台账表。返回该 Sheet 的导入结果，供 import 与 import-all 复用。
    主键交由 SQLite 自增分配；created_by/updated_at 自动填写。
    """
    result = {"sheet": sheet}

    if target_table is None:
        result["status"] = "skipped"
        result["message"] = f"Sheet '{sheet}' 无法匹配到任何台账表"
        return result

    fields = get_visible_fields_with_custom(target_table)
    field_keys = [f["key"] for f in fields]
    cn_to_key = {f["name_cn"]: f["key"] for f in fields}
    field_type_map = {f["key"]: f.get("type", "string") for f in fields}

    result["table"] = target_table
    result["table_cn_name"] = get_table_cn_name(target_table)

    # 读取表头并映射字段
    header_row = [cell.value for cell in ws[1]]
    col_mapping = {}
    for col_idx, header in enumerate(header_row, 1):
        h = str(header).strip() if header else ""
        if h in cn_to_key:
            col_mapping[col_idx] = cn_to_key[h]

    if not col_mapping:
        result["status"] = "error"
        result["message"] = f"无法匹配任何列头到「{get_table_cn_name(target_table)}」的字段"
        return result

    data_start_row = 3 if _detect_hint_row(ws) else 2
    row_errors = []
    inserted = 0
    processed = 0

    conn = get_db()
    cursor = conn.cursor()
    try:
        for row_idx in range(data_start_row, ws.max_row + 1):
            row_data = {}
            has_data = False
            for col_idx, field_key in col_mapping.items():
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    has_data = True
                    row_data[field_key] = _coerce_import_cell(cell_value)
                else:
                    row_data[field_key] = None

            if not has_data:
                continue
            processed += 1

            # 校验（仅收集错误信息，不阻止导入；skip_errors=false 时遇错即中止）
            val_warnings = []
            if do_validate:
                val_result = validate_record(target_table, row_data, db_path=DB_PATH)
                if not val_result["valid"]:
                    if skip_errors:
                        val_warnings = val_result["errors"]
                    else:
                        # 校验失败且不跳过错误：中止本表后续导入
                        row_errors.append({"row": row_idx, "errors": val_result["errors"]})
                        break

            # 入库
            try:
                filtered = {k: _coerce_value(v, field_type_map.get(k, "string"))
                            for k, v in row_data.items() if k in field_keys}
                filtered["created_by"] = username
                filtered["updated_at"] = now_beijing()

                columns = list(filtered.keys())
                placeholders = ["?"] * len(columns)
                values = [filtered[c] for c in columns]
                col_list = ", ".join('"' + c + '"' for c in columns)
                sql = (
                    f'INSERT INTO "{target_table}" '
                    f'({col_list}) '
                    f'VALUES ({", ".join(placeholders)})'
                )
                cursor.execute(sql, values)
                inserted += 1
                if val_warnings:
                    row_errors.append({"row": row_idx, "errors": val_warnings, "type": "warning"})
            except sqlite3.Error as e:
                err_msgs = [f"数据库错误: {str(e)}"]
                if val_warnings:
                    err_msgs.extend(val_warnings if isinstance(val_warnings, list) else [str(val_warnings)])
                row_errors.append({"row": row_idx, "errors": err_msgs})
                if not skip_errors:
                    break

        conn.commit()
    finally:
        conn.close()

    result["total_rows"] = processed
    result["inserted"] = inserted
    result["failed"] = processed - inserted
    result["status"] = "success" if result["failed"] == 0 else "partial"
    result["errors"] = row_errors if row_errors else None
    return result


@app.route("/api/import", methods=["POST"])
@require_admin
def import_excel():
    """
    导入 Excel 文件。
    Form data:
      - file: Excel 文件
      - table_name: 目标表名 (如果未指定, 从 Sheet 名匹配)
      - validate: 是否校验 (默认 true)
      - skip_errors: 是否跳过错误行 (默认 true)
    """
    if "file" not in request.files:
        return jsonify({"success": False, "error": "请上传 Excel 文件 (file)"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"success": False, "error": "文件名为空"}), 400

    table_name = request.form.get("table_name", "").strip()
    do_validate = request.form.get("validate", "false").lower() == "true"
    skip_errors = request.form.get("skip_errors", "true").lower() == "true"

    # 读取 Excel
    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        return jsonify({"success": False, "error": f"无法读取 Excel 文件: {str(e)}"}), 400

    results = []
    for sheet in wb.sheetnames:
        target_table = _match_sheet_to_table(sheet, table_name)
        results.append(_import_worksheet(
            wb[sheet], sheet, target_table,
            request.current_user["username"], do_validate, skip_errors,
        ))

    return jsonify({"success": True, "results": results})


# ============================================================
# API: 一键导入全部表
# ============================================================

@app.route("/api/import-all", methods=["POST"])
@require_admin
def import_all_excel():
    """
    一键导入全部表。
    接收一个多 Sheet 的 Excel 文件, 遍历所有 Sheet,
    用 Sheet 名匹配 TABLES 中的 name_cn, 对每个 Sheet 执行和 import_excel 相同的导入逻辑。
    """
    if "file" not in request.files:
        return jsonify({"success": False, "error": "请上传 Excel 文件 (file)"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"success": False, "error": "文件名为空"}), 400

    do_validate = request.form.get("validate", "false").lower() == "true"
    skip_errors = request.form.get("skip_errors", "true").lower() == "true"

    # 读取 Excel
    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        return jsonify({"success": False, "error": f"无法读取 Excel 文件: {str(e)}"}), 400

    results = []
    total_inserted = 0
    total_failed = 0
    failures = []  # 汇总所有失败记录: {sheet, row, reason}

    # 处理每个 Sheet（复用 _import_worksheet 公共导入逻辑）
    for sheet in wb.sheetnames:
        target_table = _match_sheet_to_table(sheet)
        r = _import_worksheet(
            wb[sheet], sheet, target_table,
            request.current_user["username"], do_validate, skip_errors,
        )
        results.append(r)
        total_inserted += r.get("inserted", 0)
        total_failed += r.get("failed", 0)
        if r.get("errors"):
            for err in r["errors"]:
                failures.append({
                    "sheet": sheet,
                    "row": err.get("row"),
                    "reason": "；".join(err.get("errors", [])) if isinstance(err.get("errors"), list) else str(err.get("errors")),
                })

    return jsonify({
        "success": True,
        "total_sheets": len(wb.sheetnames),
        "total_inserted": total_inserted,
        "total_failed": total_failed,
        "results": results,
        "failures": failures if failures else None,
    })


# ============================================================
# API: 导入模板下载 (R7)
# ============================================================

@app.route("/api/import-template", methods=["GET"])
@require_admin
def download_import_template():
    """生成多 Sheet 的 Excel 导入模板文件，每个 Sheet 第一行为字段中文名，第二行为格式说明"""
    wb = Workbook()
    # 删除默认 Sheet
    wb.remove(wb.active)

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hint_font = Font(name="微软雅黑", size=9, italic=True, color="666666")
    hint_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    hint_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for idx, (table_name, table_def) in enumerate(get_user_tables().items(), 1):
        sheet_name = f"{idx}.{table_def['name_cn']}"
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel Sheet 名最多31字符

        # 获取可见字段（含自定义列）作为表头
        fields = get_visible_fields_with_custom(table_name)

        # 获取合并后的下拉选项
        merged_dropdowns = _get_merged_dropdown_options(table_name, DB_PATH)

        for col_idx, field in enumerate(fields, 1):
            # 第1行: 字段中文名（表头）
            cell = ws.cell(row=1, column=col_idx, value=field["name_cn"])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

            # 第2行: 格式说明
            hints = []
            if field.get("required"):
                hints.append("必填")
            ftype = (field.get("type") or "string").lower()
            if ftype in ("number", "integer", "float", "decimal", "real"):
                hints.append("数字")
            elif ftype == "date":
                hints.append("日期: YYYY-MM-DD")
            # 下拉选项
            options = merged_dropdowns.get(field["key"]) or field.get("options")
            if options:
                hints.append("选项: " + " / ".join(options[:8]))
                if len(options) > 8:
                    hints.append(f"等{len(options)}项")

            hint_text = " | ".join(hints) if hints else "选填"
            hint_cell = ws.cell(row=2, column=col_idx, value=hint_text)
            hint_cell.font = hint_font
            hint_cell.fill = hint_fill
            hint_cell.alignment = hint_alignment
            hint_cell.border = thin_border

        # 自动列宽
        for col_idx, field in enumerate(fields, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(field["name_cn"]) + 2, 12)

        # 冻结前两行
        ws.freeze_panes = "A3"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="导入模板.xlsx",
    )


# ============================================================
# API: 批量操作
# ============================================================

@app.route("/api/<table_name>/batch", methods=["POST"])
@require_auth
@table_exists_decorator
def batch_create(table_name):
    """批量新增记录"""
    data = request.get_json(silent=True)
    if data is None or not isinstance(data, dict):
        return jsonify({"success": False, "error": "请求体必须为 JSON 格式"}), 400

    records = data.get("records", [])
    if not isinstance(records, list) or len(records) == 0:
        return jsonify({"success": False, "error": "请提供 records 列表"}), 400

    do_validate = data.get("validate", True)
    skip_errors = data.get("skip_errors", True)

    field_keys = get_all_field_keys_with_custom(table_name)
    pk = get_primary_key(table_name)

    conn = get_db()
    cursor = conn.cursor()

    inserted = 0
    failed = 0
    errors = []

    for idx, record in enumerate(records):
        record = clean_record(record)

        # 校验
        if do_validate:
            val_result = validate_record(table_name, record, db_path=DB_PATH)
            if not val_result["valid"]:
                if skip_errors:
                    failed += 1
                    errors.append({
                        "index": idx,
                        "errors": val_result["errors"],
                    })
                    continue
                else:
                    conn.close()
                    return jsonify({
                        "success": False,
                        "error": f"第 {idx + 1} 条记录校验失败",
                        "details": val_result["errors"],
                    }), 400

        try:
            # 主键交由 SQLite 自增分配（rowid/AUTOINCREMENT），避免并发 MAX(pk)+1 竞争冲突；
            # 调用方显式提供主键时随 filtered 一并插入。
            filtered = {k: v for k, v in record.items() if k in field_keys}
            # 自动记录创建者
            filtered["created_by"] = request.current_user["username"]
            # 自动记录创建时间
            filtered["updated_at"] = now_beijing()
            # 自动计算公式列初始值
            formula_values = compute_formula_fields(table_name, filtered, db_conn=conn)
            for fkey, fval in formula_values.items():
                if fkey in field_keys:
                    current_val = filtered.get(fkey)
                    if current_val is None or current_val == '':
                        filtered[fkey] = fval
            columns = list(filtered.keys())
            placeholders = ["?"] * len(columns)
            values = [filtered[c] for c in columns]

            col_list = ", ".join('"' + c + '"' for c in columns)
            sql = (
                f'INSERT INTO "{table_name}" '
                f'({col_list}) '
                f'VALUES ({", ".join(placeholders)})'
            )
            cursor.execute(sql, values)
            inserted += 1
        except sqlite3.Error as e:
            failed += 1
            errors.append({
                "index": idx,
                "errors": [f"数据库错误: {str(e)}"],
            })

    conn.commit()
    conn.close()

    return jsonify({
        "success": True,
        "table_name": table_name,
        "table_cn_name": get_table_cn_name(table_name),
        "total": len(records),
        "inserted": inserted,
        "failed": failed,
        "errors": errors if errors else None,
    })


# ============================================================
# API: 统计信息
# ============================================================

@app.route("/api/stats", methods=["GET"])
@require_auth
def get_stats():
    """获取所有表的记录数统计（仅统计用户可见表）"""
    conn = get_db()
    cursor = conn.cursor()
    user_tables = get_user_tables()
    stats = []
    for table_name, table_info in user_tables.items():
        try:
            cursor.execute(f'SELECT COUNT(*) FROM "{table_name}"')
            count = cursor.fetchone()[0]
        except sqlite3.Error:
            count = 0
        stats.append({
            "table_name": table_name,
            "name_cn": table_info["name_cn"],
            "record_count": count,
            "field_count": len(table_info["fields"]),
        })
    conn.close()

    total_records = sum(s["record_count"] for s in stats)
    return jsonify({
        "success": True,
        "total_tables": len(user_tables),
        "total_records": total_records,
        "stats": stats,
    })


# ============================================================
# API: 健康检查
# ============================================================

@app.route("/api/health", methods=["GET"])
def health_check():
    """健康检查（不暴露内部文件路径）"""
    return jsonify({
        "success": True,
        "status": "healthy",
        "service": "采购代理台账管理系统",
        "version": "2.0.0",
        "tables": list(get_user_tables().keys()),
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })


# ============================================================
# 错误处理
# ============================================================

@app.errorhandler(404)
def not_found(e):
    return jsonify({
        "success": False,
        "error": "接口不存在",
        "path": request.path,
    }), 404


@app.errorhandler(405)
def method_not_allowed(e):
    return jsonify({
        "success": False,
        "error": f"不支持的请求方法: {request.method}",
        "path": request.path,
    }), 405


@app.errorhandler(500)
def internal_error(e):
    # 生产环境不暴露内部异常详情，仅调试模式返回
    detail = str(e) if app.config.get("DEBUG") else "请稍后重试或联系管理员"
    return jsonify({
        "success": False,
        "error": "服务器内部错误",
        "detail": detail,
    }), 500


# ============================================================
# API: 数据分析
# ============================================================

# ============================================================
# API: 清空台账
# ============================================================

@app.route("/api/clear-all", methods=["POST"])
@require_admin
def clear_all_data():
    """清空所有业务台账数据（仅管理员）。

    前置条件: 前端在调用此接口前必须强制用户先导出全部数据。
    """
    # 需要确认参数（兼容 JSON 与 form 两种请求方式）
    body = request.get_json(silent=True) or {}
    confirm = str(body.get("confirm") or request.form.get("confirm", "")).strip()
    if confirm != "DELETE_ALL_DATA":
        return jsonify({
            "success": False,
            "error": "缺少确认参数，请通过前端界面操作",
        }), 400

    conn = get_db()
    cursor = conn.cursor()

    cleared = []
    total_deleted = 0

    for table_name in get_user_tables():
        try:
            cursor.execute(f'SELECT COUNT(*) FROM "{table_name}"')
            count = cursor.fetchone()[0]
            if count > 0:
                cursor.execute(f'DELETE FROM "{table_name}"')
                cleared.append({"table": table_name, "deleted": count})
                total_deleted += count
        except sqlite3.Error as e:
            cleared.append({"table": table_name, "error": str(e)})

    # 重置所有表的自增 ID
    for table_name in get_user_tables():
        try:
            cursor.execute(
                "UPDATE sqlite_sequence SET seq = 0 "
                "WHERE name = ?",
                (table_name,),
            )
        except sqlite3.Error:
            pass  # 表可能没有自增序列

    conn.commit()
    conn.close()

    return jsonify({
        "success": True,
        "message": f"已清空全部台账数据，共删除 {total_deleted} 条记录",
        "cleared_tables": cleared,
        "total_deleted": total_deleted,
    })


@app.route("/analytics")
def analytics_page():
    """数据分析页面（权限由前端JS + API层的 @require_admin 控制）"""
    return render_template("analytics.html")


def _safe_float(val):
    """安全转换为 float，失败返回 0。"""
    try:
        if val is None or val == "":
            return 0.0
        return float(str(val).replace(",", "").replace("，", ""))
    except (ValueError, TypeError):
        return 0.0


@app.route("/api/analytics/overview", methods=["GET"])
@require_admin
def analytics_overview():
    """总览统计：各表记录数、项目金额汇总、超期统计"""
    conn = get_db()
    cursor = conn.cursor()

    # 1. 各表记录数
    table_counts = {}
    for t_name in get_user_tables():
        try:
            cursor.execute(f'SELECT COUNT(*) FROM "{t_name}"')
            table_counts[t_name] = cursor.fetchone()[0]
        except sqlite3.Error:
            table_counts[t_name] = 0

    # 2. 项目台账金额汇总
    cursor.execute(
        'SELECT '
        '  COUNT(*), '
        '  COALESCE(SUM(CASE WHEN budget_amount IS NOT NULL AND budget_amount != "" THEN CAST(budget_amount AS REAL) ELSE 0 END), 0), '
        '  COALESCE(SUM(CASE WHEN winning_amount IS NOT NULL AND winning_amount != "" THEN CAST(winning_amount AS REAL) ELSE 0 END), 0), '
        '  COALESCE(SUM(CASE WHEN agency_fee_amount IS NOT NULL AND agency_fee_amount != "" THEN CAST(agency_fee_amount AS REAL) ELSE 0 END), 0) '
        'FROM projects'
    )
    row = cursor.fetchone()
    project_stats = {
        "total": row[0],
        "total_budget": round(row[1], 2),
        "total_winning": round(row[2], 2),
        "total_agency_fee": round(row[3], 2),
    }

    # 3. 项目按采购环节分布
    cursor.execute(
        'SELECT COALESCE(NULLIF(current_stage, ""), "未填写"), COUNT(*) '
        'FROM projects GROUP BY current_stage ORDER BY COUNT(*) DESC'
    )
    stage_dist = [{"name": r[0], "value": r[1]} for r in cursor.fetchall()]

    # 4. 项目按采购方式分布
    cursor.execute(
        'SELECT COALESCE(NULLIF(purchase_method, ""), "未填写"), COUNT(*) '
        'FROM projects GROUP BY purchase_method ORDER BY COUNT(*) DESC'
    )
    method_dist = [{"name": r[0], "value": r[1]} for r in cursor.fetchall()]

    # 5. 项目按合同期分布
    cursor.execute(
        'SELECT COALESCE(NULLIF(contract_period, ""), "未填写"), COUNT(*) '
        'FROM projects GROUP BY contract_period ORDER BY contract_period'
    )
    period_dist = [{"name": r[0], "value": r[1]} for r in cursor.fetchall()]

    # 6. 超期统计
    cursor.execute(
        'SELECT '
        '  SUM(CASE WHEN record_is_overdue = "是" THEN 1 ELSE 0 END), '
        '  SUM(CASE WHEN impl_report_is_overdue = "是" THEN 1 ELSE 0 END), '
        '  SUM(CASE WHEN archive_is_overdue = "是" THEN 1 ELSE 0 END) '
        'FROM projects'
    )
    o_row = cursor.fetchone()
    overdue_stats = {
        "record_overdue": o_row[0] or 0,
        "impl_report_overdue": o_row[1] or 0,
        "archive_overdue": o_row[2] or 0,
    }

    # 7. 问题统计
    cursor.execute(
        'SELECT COALESCE(NULLIF(issue_type, ""), "未分类"), COUNT(*) '
        'FROM project_issues GROUP BY issue_type ORDER BY COUNT(*) DESC'
    )
    issue_dist = [{"name": r[0], "value": r[1]} for r in cursor.fetchall()]

    cursor.execute(
        'SELECT '
        '  COUNT(*), '
        '  SUM(CASE WHEN is_rectified = "是" THEN 1 ELSE 0 END) '
        'FROM project_issues'
    )
    i_row = cursor.fetchone()
    issue_stats = {"total": i_row[0] or 0, "rectified": i_row[1] or 0}

    conn.close()

    return jsonify({
        "success": True,
        "table_counts": table_counts,
        "project_stats": project_stats,
        "stage_distribution": stage_dist,
        "method_distribution": method_dist,
        "period_distribution": period_dist,
        "overdue_stats": overdue_stats,
        "issue_distribution": issue_dist,
        "issue_stats": issue_stats,
    })


@app.route("/api/analytics/projects", methods=["GET"])
@require_admin
def analytics_projects():
    """项目台账深度分析"""
    conn = get_db()
    cursor = conn.cursor()

    # 1. 按项目类别分布
    cursor.execute(
        'SELECT COALESCE(NULLIF(project_category, ""), "未填写"), COUNT(*) '
        'FROM projects GROUP BY project_category ORDER BY COUNT(*) DESC'
    )
    category_dist = [{"name": r[0], "value": r[1]} for r in cursor.fetchall()]

    # 2. 按委托单位 Top 10
    cursor.execute(
        'SELECT COALESCE(NULLIF(client_name, ""), "未填写"), COUNT(*) as cnt '
        'FROM projects GROUP BY client_name ORDER BY cnt DESC LIMIT 10'
    )
    client_dist = [{"name": r[0], "value": r[1]} for r in cursor.fetchall()]

    # 3. 按采购经理分布
    cursor.execute(
        'SELECT COALESCE(NULLIF(purchase_manager, ""), "未填写"), COUNT(*) as cnt '
        'FROM projects GROUP BY purchase_manager ORDER BY cnt DESC'
    )
    manager_dist = [{"name": r[0], "value": r[1]} for r in cursor.fetchall()]

    # 4. 按合同期统计预算/中标/代理费金额
    cursor.execute(
        'SELECT '
        '  COALESCE(NULLIF(contract_period, ""), "未填写"), '
        '  COUNT(*), '
        '  COALESCE(SUM(CASE WHEN budget_amount IS NOT NULL AND budget_amount != "" THEN CAST(budget_amount AS REAL) ELSE 0 END), 0), '
        '  COALESCE(SUM(CASE WHEN winning_amount IS NOT NULL AND winning_amount != "" THEN CAST(winning_amount AS REAL) ELSE 0 END), 0), '
        '  COALESCE(SUM(CASE WHEN agency_fee_amount IS NOT NULL AND agency_fee_amount != "" THEN CAST(agency_fee_amount AS REAL) ELSE 0 END), 0) '
        'FROM projects GROUP BY contract_period ORDER BY contract_period'
    )
    period_amount = [{
        "period": r[0],
        "count": r[1],
        "budget": round(r[2], 2),
        "winning": round(r[3], 2),
        "agency_fee": round(r[4], 2),
    } for r in cursor.fetchall()]

    # 5. 归档状态
    cursor.execute(
        'SELECT COALESCE(NULLIF(is_archived, ""), "未填写"), COUNT(*) '
        'FROM projects GROUP BY is_archived ORDER BY COUNT(*) DESC'
    )
    archive_dist = [{"name": r[0], "value": r[1]} for r in cursor.fetchall()]

    # 6. 稽核问题
    cursor.execute(
        'SELECT COALESCE(NULLIF(audit_has_issues, ""), "未填写"), COUNT(*) '
        'FROM projects GROUP BY audit_has_issues'
    )
    audit_dist = [{"name": r[0], "value": r[1]} for r in cursor.fetchall()]

    conn.close()

    return jsonify({
        "success": True,
        "category_distribution": category_dist,
        "client_top10": client_dist,
        "manager_distribution": manager_dist,
        "period_amount": period_amount,
        "archive_distribution": archive_dist,
        "audit_distribution": audit_dist,
    })


@app.route("/api/analytics/financial", methods=["GET"])
@require_admin
def analytics_financial():
    """财务分析：代理费、保证金、评审费"""
    conn = get_db()
    cursor = conn.cursor()

    # 1. 代理费台账汇总
    cursor.execute(
        'SELECT '
        '  COUNT(*), '
        '  COALESCE(SUM(CASE WHEN receivable_total IS NOT NULL AND receivable_total != "" THEN CAST(receivable_total AS REAL) ELSE 0 END), 0), '
        '  COALESCE(SUM(CASE WHEN actual_received_fee IS NOT NULL AND actual_received_fee != "" THEN CAST(actual_received_fee AS REAL) ELSE 0 END), 0) '
        'FROM agency_fees'
    )
    row = cursor.fetchone()
    agency_fee_stats = {
        "total_records": row[0],
        "total_receivable": round(row[1], 2),
        "total_received": round(row[2], 2),
        "total_outstanding": round(row[1] - row[2], 2),
    }

    # 2. 代理费按合同期
    cursor.execute(
        'SELECT '
        '  COALESCE(NULLIF(contract_period, ""), "未填写"), '
        '  COUNT(*), '
        '  COALESCE(SUM(CASE WHEN receivable_total IS NOT NULL AND receivable_total != "" THEN CAST(receivable_total AS REAL) ELSE 0 END), 0), '
        '  COALESCE(SUM(CASE WHEN actual_received_fee IS NOT NULL AND actual_received_fee != "" THEN CAST(actual_received_fee AS REAL) ELSE 0 END), 0) '
        'FROM agency_fees GROUP BY contract_period ORDER BY contract_period'
    )
    agency_by_period = [{
        "period": r[0],
        "count": r[1],
        "receivable": round(r[2], 2),
        "received": round(r[3], 2),
    } for r in cursor.fetchall()]

    # 3. 投标保证金汇总
    cursor.execute(
        'SELECT '
        '  COUNT(*), '
        '  COALESCE(SUM(CASE WHEN receivable_amount IS NOT NULL AND receivable_amount != "" THEN CAST(receivable_amount AS REAL) ELSE 0 END), 0), '
        '  COALESCE(SUM(CASE WHEN received_amount IS NOT NULL AND received_amount != "" THEN CAST(received_amount AS REAL) ELSE 0 END), 0), '
        '  COALESCE(SUM(CASE WHEN actual_refund_amount IS NOT NULL AND actual_refund_amount != "" THEN CAST(actual_refund_amount AS REAL) ELSE 0 END), 0) '
        'FROM bid_deposits'
    )
    row = cursor.fetchone()
    bid_deposit_stats = {
        "total_records": row[0],
        "total_receivable": round(row[1], 2),
        "total_received": round(row[2], 2),
        "total_refunded": round(row[3], 2),
    }

    # 4. 履约保证金汇总
    cursor.execute(
        'SELECT '
        '  COUNT(*), '
        '  COALESCE(SUM(CASE WHEN deposit_amount IS NOT NULL AND deposit_amount != "" THEN CAST(deposit_amount AS REAL) ELSE 0 END), 0) '
        'FROM performance_deposits'
    )
    row = cursor.fetchone()
    perf_deposit_stats = {
        "total_records": row[0],
        "total_amount": round(row[1], 2),
    }

    # 5. 评审费汇总
    cursor.execute(
        'SELECT '
        '  COUNT(*), '
        '  COALESCE(SUM(CASE WHEN total_fee IS NOT NULL AND total_fee != "" THEN CAST(total_fee AS REAL) ELSE 0 END), 0) '
        'FROM expert_fees'
    )
    row = cursor.fetchone()
    expert_fee_stats = {
        "total_records": row[0],
        "total_fee": round(row[1], 2),
    }

    # 6. 评审费按发放方式
    cursor.execute(
        'SELECT COALESCE(NULLIF(payment_method, ""), "未填写"), COUNT(*), '
        '  COALESCE(SUM(CASE WHEN total_fee IS NOT NULL AND total_fee != "" THEN CAST(total_fee AS REAL) ELSE 0 END), 0) '
        'FROM expert_fees GROUP BY payment_method ORDER BY COUNT(*) DESC'
    )
    expert_by_method = [{
        "name": r[0],
        "count": r[1],
        "amount": round(r[2], 2),
    } for r in cursor.fetchall()]

    conn.close()

    return jsonify({
        "success": True,
        "agency_fee_stats": agency_fee_stats,
        "agency_by_period": agency_by_period,
        "bid_deposit_stats": bid_deposit_stats,
        "perf_deposit_stats": perf_deposit_stats,
        "expert_fee_stats": expert_fee_stats,
        "expert_by_method": expert_by_method,
    })


@app.route("/api/analytics/personnel", methods=["GET"])
@require_admin
def analytics_personnel():
    """人员分析"""
    conn = get_db()
    cursor = conn.cursor()

    # 1. 人员总数和状态
    cursor.execute(
        'SELECT '
        '  COUNT(*), '
        '  SUM(CASE WHEN status = "正常" THEN 1 ELSE 0 END), '
        '  SUM(CASE WHEN status = "冻结" THEN 1 ELSE 0 END), '
        '  SUM(CASE WHEN status = "锁定" THEN 1 ELSE 0 END), '
        '  SUM(CASE WHEN is_resigned = "是" THEN 1 ELSE 0 END) '
        'FROM personnel'
    )
    row = cursor.fetchone()
    personnel_stats = {
        "total": row[0] or 0,
        "normal": row[1] or 0,
        "frozen": row[2] or 0,
        "locked": row[3] or 0,
        "resigned": row[4] or 0,
    }

    # 2. 按代理机构分布
    cursor.execute(
        'SELECT COALESCE(NULLIF(agency_name, ""), "未填写"), COUNT(*) as cnt '
        'FROM personnel GROUP BY agency_name ORDER BY cnt DESC'
    )
    agency_dist = [{"name": r[0], "value": r[1]} for r in cursor.fetchall()]

    # 3. 按资质分布
    cursor.execute(
        'SELECT COALESCE(NULLIF(procurement_qualification, ""), "未填写"), COUNT(*) '
        'FROM personnel GROUP BY procurement_qualification'
    )
    qual_dist = [{"name": r[0], "value": r[1]} for r in cursor.fetchall()]

    # 4. 认证过期统计
    cursor.execute(
        'SELECT COALESCE(NULLIF(cert_is_expired, ""), "未填写"), COUNT(*) '
        'FROM personnel GROUP BY cert_is_expired'
    )
    cert_dist = [{"name": r[0], "value": r[1]} for r in cursor.fetchall()]

    conn.close()

    return jsonify({
        "success": True,
        "personnel_stats": personnel_stats,
        "agency_distribution": agency_dist,
        "qualification_distribution": qual_dist,
        "cert_distribution": cert_dist,
    })


# ============================================================
# 主入口
# ============================================================

if __name__ == "__main__":
    # 启动前确保表已创建
    ensure_tables()

    print("=" * 60)
    print("  采购代理台账管理系统 - Flask 后端")
    print("=" * 60)
    print(f"  数据库路径: {DB_PATH}")
    print(f"  台账表数量: {len(get_user_tables())}")
    print(f"  监听地址:   0.0.0.0:5000")
    print(f"  调试模式:   {'开启 (FLASK_DEBUG)' if app.config['DEBUG'] else '关闭'}")
    print("  默认管理员: admin / admin123")
    if app.config.get("DEFAULT_PASSWORD_WARNING"):
        print("  " + "!" * 56)
        print("  ! 警告: 默认管理员仍使用默认密码 admin/admin123")
        print("  !       请登录后立即通过「用户管理」修改密码")
        print("  !       部署到公网/多人环境前必须修改")
        print("  " + "!" * 56)
    print("=" * 60)
    print("  可用 API:")
    print("  --- 认证相关 ---")
    print("  POST   /api/login              - 用户登录, 返回 token")
    print("  POST   /api/logout             - 用户登出")
    print("  GET    /api/me                 - 获取当前登录用户")
    print("  --- 用户管理 (仅管理员) ---")
    print("  GET    /api/users              - 获取所有用户列表")
    print("  POST   /api/users              - 新增用户")
    print("  PUT    /api/users/<id>         - 修改用户")
    print("  DELETE /api/users/<id>         - 删除用户")
    print("  --- 台账数据 (需登录) ---")
    print("  GET    /api/tables              - 获取所有表信息")
    print("  GET    /api/columns/<table>     - 获取字段定义和下拉菜单")
    print("  GET    /api/columns/all/<table> - 获取全部列(含隐藏/自定义)")
    print("  POST   /api/columns/hide        - 隐藏基础列 (管理员)")
    print("  POST   /api/columns/restore     - 恢复基础列 (管理员)")
    print("  POST   /api/columns/rename      - 重命名表头列 (管理员)")
    print("  DELETE /api/columns/rename      - 恢复表头原始名称 (管理员)")
    print("  GET    /api/dropdowns          - 获取所有下拉选项")
    print("  GET    /api/dropdowns/<table>   - 获取指定表下拉选项")
    print("  POST   /api/dropdowns           - 添加下拉选项 (管理员)")
    print("  DELETE /api/dropdowns/<id>      - 删除下拉选项 (管理员)")
    print("  POST   /api/dropdowns/enable    - 启用字段下拉功能 (管理员)")
    print("  GET    /api/custom-columns      - 获取所有自定义列")
    print("  POST   /api/custom-columns      - 添加自定义列 (管理员)")
    print("  DELETE /api/custom-columns/<id> - 删除自定义列 (管理员)")
    print("  GET    /api/<table>             - 获取记录列表 (支持分页/搜索)")
    print("  GET    /api/<table>/<id>        - 获取单条记录")
    print("  POST   /api/<table>             - 新增记录")
    print("  PUT    /api/<table>/<id>        - 更新记录")
    print("  DELETE /api/<table>/<id>        - 删除记录")
    print("  POST   /api/<table>/validate    - 校验记录")
    print("  POST   /api/<table>/batch       - 批量新增")
    print("  GET    /api/export/<table>     - 导出 Excel")
    print("  GET    /api/export-all          - 导出所有表 Excel (管理员)")
    print("  POST   /api/import             - 导入 Excel (管理员)")
    print("  POST   /api/import-all         - 一键导入全部表 (管理员)")
    print("  GET    /api/import-template    - 下载导入模板 (管理员)")
    print("  GET    /api/stats              - 统计信息")
    print("  GET    /api/health             - 健康检查")
    print("=" * 60)

    # debug 模式由 FLASK_DEBUG 环境变量控制，默认关闭（生产安全）
    app.run(host="0.0.0.0", port=5000, debug=app.config["DEBUG"])
